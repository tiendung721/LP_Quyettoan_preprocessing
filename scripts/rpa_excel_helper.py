"""Helper CLI for PAD to mark one Excel information row as imported."""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
import uuid
from copy import copy
from datetime import datetime
from enum import IntEnum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

COMMAND_FINISH_INFO = "finish-info"
HEADER_SQT = "SQT PM"
HEADER_STATUS = "Trạng thái nhập"
HEADER_LAST_IMPORTED = "Ngày nhập cuối"
STATUS_IMPORTED = "Đã nhập"
DISPLAY_DATETIME_FORMAT = "%d/%m/%Y %H:%M:%S"
EXCEL_DATETIME_FORMAT = "dd/mm/yyyy hh:mm:ss"
MAX_SAVE_ATTEMPTS = 3
SAVE_RETRY_SECONDS = 1


class ExitCode(IntEnum):
    SUCCESS = 0
    UNKNOWN_ERROR = 1
    INVALID_ARGUMENTS = 2
    FILE_NOT_FOUND = 3
    SHEET_NOT_FOUND = 4
    INVALID_ROW = 5
    REQUIRED_COLUMN_NOT_FOUND = 6
    FILE_LOCKED_OR_SAVE_FAILED = 7
    WORKBOOK_READ_ERROR = 8


class ExcelHelperError(RuntimeError):
    """Typed error that maps to an exit code for PAD."""

    def __init__(self, exit_code: ExitCode, error_code: str, message: str):
        super().__init__(message)
        self.exit_code = exit_code
        self.error_code = error_code
        self.message = message


class CliArgumentError(ValueError):
    pass


class JsonArgumentParser(argparse.ArgumentParser):
    def error(self, message: str) -> None:
        raise CliArgumentError(message)


def project_root() -> Path:
    return Path(__file__).resolve().parents[1]


def setup_logging() -> logging.Logger:
    """Configure a file-only logger so stdout remains valid JSON for PAD."""
    logs_dir = project_root() / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("rpa_excel_helper")
    logger.setLevel(logging.INFO)
    logger.propagate = False
    if not logger.handlers:
        handler = logging.FileHandler(logs_dir / "rpa_excel_helper.log", encoding="utf-8")
        handler.setFormatter(
            logging.Formatter(
                "%(asctime)s [%(levelname)s] %(message)s",
                "%Y-%m-%d %H:%M:%S",
            )
        )
        logger.addHandler(handler)
    return logger


def build_parser() -> argparse.ArgumentParser:
    """Build the command-line parser used by PAD."""
    parser = JsonArgumentParser(
        prog="rpa_excel_helper.py",
        description="Cập nhật trạng thái nhập cho đúng một dòng Excel.",
    )
    subparsers = parser.add_subparsers(
        dest="command",
        required=True,
        parser_class=JsonArgumentParser,
    )

    finish = subparsers.add_parser(COMMAND_FINISH_INFO)
    finish.add_argument("--file", required=True, help="Đường dẫn file Excel hằng ngày.")
    finish.add_argument("--sheet", required=True, help="Tên sheet cần cập nhật.")
    finish.add_argument("--row", required=True, help="Số dòng vật lý trong Excel.")
    return parser


def normalize_header(value: object) -> str:
    """Normalize a header for exact trimmed-name comparison."""
    return "" if value is None else str(value).strip()


def find_header_columns(worksheet) -> dict[str, int]:
    """Return the first column index for each non-empty header in row 1."""
    columns: dict[str, int] = {}
    for cell in worksheet[1]:
        header = normalize_header(cell.value)
        if header and header not in columns:
            columns[header] = cell.column
    return columns


def ensure_last_import_column(worksheet, status_column: int) -> int:
    """Find or add the 'Ngày nhập cuối' column exactly once."""
    columns = find_header_columns(worksheet)
    if HEADER_LAST_IMPORTED in columns:
        return columns[HEADER_LAST_IMPORTED]

    new_column = worksheet.max_column + 1
    source = worksheet.cell(1, status_column)
    target = worksheet.cell(1, new_column)
    target.value = HEADER_LAST_IMPORTED
    if source.has_style:
        target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    worksheet.column_dimensions[get_column_letter(new_column)].width = 24
    return new_column


def parse_row_number(value: str) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise ExcelHelperError(
            ExitCode.INVALID_ROW,
            "INVALID_ROW",
            f"Số dòng Excel không hợp lệ: {value!r}.",
        ) from exc


def validate_excel_row(worksheet, row_number: int, sqt_column: int) -> str:
    """Validate the physical Excel row and return its SQT value."""
    if row_number <= 1:
        raise ExcelHelperError(
            ExitCode.INVALID_ROW,
            "INVALID_ROW",
            "Số dòng Excel phải lớn hơn 1 vì dòng 1 là header.",
        )
    if row_number > worksheet.max_row:
        raise ExcelHelperError(
            ExitCode.INVALID_ROW,
            "INVALID_ROW",
            f"Số dòng Excel {row_number} vượt quá số dòng hiện có ({worksheet.max_row}).",
        )

    row_values = [
        worksheet.cell(row_number, col).value
        for col in range(1, worksheet.max_column + 1)
    ]
    if not any(value not in (None, "") for value in row_values):
        raise ExcelHelperError(
            ExitCode.INVALID_ROW,
            "EMPTY_ROW",
            f"Dòng Excel {row_number} đang trống, không cập nhật trạng thái nhập.",
        )

    sqt = worksheet.cell(row_number, sqt_column).value
    sqt_text = "" if sqt is None else str(sqt).strip()
    if not sqt_text:
        raise ExcelHelperError(
            ExitCode.INVALID_ROW,
            "EMPTY_SQT",
            f"Dòng Excel {row_number} không có giá trị SQT PM.",
        )
    if sqt_text.endswith(".0") and sqt_text[:-2].isdigit():
        sqt_text = sqt_text[:-2]
    return sqt_text


def _remove_temp_file(path: Path) -> None:
    try:
        if path.exists():
            path.unlink()
    except OSError:
        pass


def atomic_save_workbook(workbook, destination: Path, logger: logging.Logger) -> None:
    """Save to a temporary workbook first, then replace the original file."""
    suffix = destination.suffix or ".xlsx"
    temp_path = destination.with_name(
        f".{destination.name}.rpa_tmp_{uuid.uuid4().hex}{suffix}"
    )
    save_error: Exception | None = None
    try:
        for attempt in range(1, MAX_SAVE_ATTEMPTS + 1):
            try:
                workbook.save(temp_path)
                save_error = None
                break
            except (PermissionError, OSError) as exc:
                save_error = exc
                logger.warning(
                    "Lưu file tạm thất bại lần %s/%s: %s",
                    attempt,
                    MAX_SAVE_ATTEMPTS,
                    exc,
                )
                if attempt < MAX_SAVE_ATTEMPTS:
                    time.sleep(SAVE_RETRY_SECONDS)
        if save_error is not None:
            raise ExcelHelperError(
                ExitCode.FILE_LOCKED_OR_SAVE_FAILED,
                "SAVE_FAILED",
                "Không thể cập nhật file Excel vì file đang được mở hoặc bị khóa. "
                "Vui lòng đóng Excel và thử lại.",
            ) from save_error

        workbook.close()
        replace_error: Exception | None = None
        for attempt in range(1, MAX_SAVE_ATTEMPTS + 1):
            try:
                os.replace(temp_path, destination)
                replace_error = None
                return
            except (PermissionError, OSError) as exc:
                replace_error = exc
                logger.warning(
                    "Thay thế file gốc thất bại lần %s/%s: %s",
                    attempt,
                    MAX_SAVE_ATTEMPTS,
                    exc,
                )
                if attempt < MAX_SAVE_ATTEMPTS:
                    time.sleep(SAVE_RETRY_SECONDS)
        if replace_error is not None:
            raise ExcelHelperError(
                ExitCode.FILE_LOCKED_OR_SAVE_FAILED,
                "SAVE_FAILED",
                "Không thể cập nhật file Excel vì file đang được mở hoặc bị khóa. "
                "Vui lòng đóng Excel và thử lại.",
            ) from replace_error
    finally:
        _remove_temp_file(temp_path)


def update_information_status(
    file_path: Path,
    sheet_name: str,
    row_number: int,
    logger: logging.Logger | None = None,
) -> dict[str, Any]:
    """Mark exactly one information row as imported after PAD confirms web save."""
    logger = logger or setup_logging()
    destination = file_path.resolve()
    if not destination.is_file():
        raise ExcelHelperError(
            ExitCode.FILE_NOT_FOUND,
            "FILE_NOT_FOUND",
            f"Không tìm thấy file Excel: {destination}",
        )

    workbook = None
    saved = False
    try:
        try:
            workbook = load_workbook(destination)
        except PermissionError as exc:
            raise ExcelHelperError(
                ExitCode.FILE_LOCKED_OR_SAVE_FAILED,
                "FILE_LOCKED",
                "Không thể mở file Excel vì file đang được mở hoặc bị khóa. "
                "Vui lòng đóng Excel và thử lại.",
            ) from exc
        except (InvalidFileException, OSError, ValueError) as exc:
            raise ExcelHelperError(
                ExitCode.WORKBOOK_READ_ERROR,
                "WORKBOOK_READ_ERROR",
                f"Không đọc được workbook Excel: {destination}",
            ) from exc

        if sheet_name not in workbook.sheetnames:
            raise ExcelHelperError(
                ExitCode.SHEET_NOT_FOUND,
                "SHEET_NOT_FOUND",
                f"Không tìm thấy sheet {sheet_name}.",
            )
        worksheet = workbook[sheet_name]
        columns = find_header_columns(worksheet)
        missing = [
            header
            for header in (HEADER_SQT, HEADER_STATUS)
            if header not in columns
        ]
        if missing:
            raise ExcelHelperError(
                ExitCode.REQUIRED_COLUMN_NOT_FOUND,
                "REQUIRED_COLUMN_NOT_FOUND",
                "Không tìm thấy cột bắt buộc: " + ", ".join(missing),
            )

        status_column = columns[HEADER_STATUS]
        last_import_column = ensure_last_import_column(worksheet, status_column)
        sqt = validate_excel_row(worksheet, row_number, columns[HEADER_SQT])
        previous_status = worksheet.cell(row_number, status_column).value
        updated_at = datetime.now().strftime(DISPLAY_DATETIME_FORMAT)

        worksheet.cell(row_number, status_column).value = STATUS_IMPORTED
        last_import_cell = worksheet.cell(row_number, last_import_column)
        last_import_cell.value = updated_at
        last_import_cell.number_format = EXCEL_DATETIME_FORMAT

        logger.info(
            "finish-info file=%s sheet=%s row=%s sqt=%s status_before=%s status_after=%s updated_at=%s",
            destination,
            sheet_name,
            row_number,
            sqt,
            previous_status,
            STATUS_IMPORTED,
            updated_at,
        )
        atomic_save_workbook(workbook, destination, logger)
        saved = True
        return {
            "success": True,
            "command": COMMAND_FINISH_INFO,
            "file": str(destination),
            "sheet": sheet_name,
            "row": row_number,
            "sqt": sqt,
            "status": STATUS_IMPORTED,
            "updated_at": updated_at,
        }
    finally:
        if workbook is not None and not saved:
            try:
                workbook.close()
            except Exception:
                pass


def write_stdout_json(payload: dict[str, Any]) -> None:
    data = (json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n").encode(
        "utf-8"
    )
    sys.stdout.buffer.write(data)
    sys.stdout.flush()


def write_stderr_json(payload: dict[str, Any]) -> None:
    data = (json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n").encode(
        "utf-8"
    )
    sys.stderr.buffer.write(data)
    sys.stderr.flush()


def handle_finish_info(args: argparse.Namespace, logger: logging.Logger) -> int:
    row_number = parse_row_number(args.row)
    result = update_information_status(
        Path(args.file),
        str(args.sheet),
        row_number,
        logger,
    )
    write_stdout_json(result)
    return int(ExitCode.SUCCESS)


def main(argv: list[str] | None = None) -> int:
    logger = setup_logging()
    parser = build_parser()
    try:
        args = parser.parse_args(argv)
        logger.info(
            "Chạy command=%s file=%s sheet=%s row=%s",
            getattr(args, "command", ""),
            getattr(args, "file", ""),
            getattr(args, "sheet", ""),
            getattr(args, "row", ""),
        )
        if args.command == COMMAND_FINISH_INFO:
            return handle_finish_info(args, logger)
        raise ExcelHelperError(
            ExitCode.INVALID_ARGUMENTS,
            "INVALID_COMMAND",
            f"Command không được hỗ trợ: {args.command}",
        )
    except CliArgumentError as exc:
        payload = {
            "success": False,
            "error_code": "INVALID_ARGUMENTS",
            "message": f"Tham số dòng lệnh không hợp lệ: {exc}",
        }
        logger.error(payload["message"])
        write_stderr_json(payload)
        return int(ExitCode.INVALID_ARGUMENTS)
    except ExcelHelperError as exc:
        payload = {
            "success": False,
            "error_code": exc.error_code,
            "message": exc.message,
        }
        logger.error("%s: %s", exc.error_code, exc.message)
        write_stderr_json(payload)
        return int(exc.exit_code)
    except Exception as exc:  # noqa: BLE001 - PAD needs a non-zero exit code
        logger.exception("Lỗi không xác định khi chạy rpa_excel_helper.")
        write_stderr_json(
            {
                "success": False,
                "error_code": "UNKNOWN_ERROR",
                "message": f"Lỗi không xác định: {exc}",
            }
        )
        return int(ExitCode.UNKNOWN_ERROR)


if __name__ == "__main__":
    raise SystemExit(main())
