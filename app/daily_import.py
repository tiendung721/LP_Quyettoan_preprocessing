"""Phân tích và nhập output GPT vào file quyết toán hằng ngày.

Module này không phụ thuộc giao diện Qt. Giao diện chỉ gọi ``analyze`` để lấy
kế hoạch, thu thập lựa chọn của người dùng rồi gọi ``commit`` để ghi an toàn.

File JSON bóc tách được coi là bộ nhớ tạm: mỗi lần phân tích chỉ làm việc với
đúng các dòng đang có trong file, không kéo dữ liệu chờ của các lần trước vào.
Dòng nào không ghép được với dữ liệu quyết toán thì trả về trong
``ImportAnalysis.unmatched_rows`` để người dùng sửa lại ở Bước 2 hoặc bỏ đi,
chứ không ghi nửa vời vào file theo dõi.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import unicodedata
from copy import copy
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import file_utils
from .database import Database


INFO_SHEET = "Thong_Tin_Quyet_Toan"
EXPENSE_SHEET = "Khoan_Chi"

INFO_HEADERS = [
    "SQT PM",
    "Ngày Đóng",
    "Số Container",
    "Biển số xe",
    "Số chì/Seal",
    "Lô SX",
    "Số tấn",
    "Loại hàng",
    "Nơi đóng",
    "Tên tàu",
    "Ngày chạy",
    "Dự kiến giao",
    "Người nhận",
    "VT biển",
    "Vận chuyển",
    "HD HP",
    "HD HCM",
    "Ngày Giao",
    "Hóa Đơn quyết toán",
    "Ghi chú",
    "Trạng thái nhập",
    "Ngày nhập cuối",
    "Ngày cập nhật",
    "MD5",
]

EXPENSE_HEADERS = [
    "SQT PM",
    "Ngày tháng",
    "Số Container",
    "Số HĐ",
    "Giá vật liệu",
    "Đơn giá",
    "Thành tiền",
    "VAT",
    "Tổng tiền",
    "Trạng thái nhập",
    "Ngày cập nhật",
    "MD5",
]

MD5_HEADER_ALIASES = ("MD5", "Mã MD5")

DOC_SCALE = "PHIEU_CAN"
DOC_BILL = "BILL"
DOC_EXPENSE = "KHOAN_CHI"

DOC_LABELS = {
    DOC_SCALE: "Phiếu cân",
    DOC_BILL: "Bill",
    DOC_EXPENSE: "Khoản chi",
}

STATE_READY = "READY"
STATE_COMPLETED = "COMPLETED"
STATE_IGNORED = "IGNORED"

# Lý do một dòng bóc tách không ghép được với dữ liệu quyết toán.
REASON_NO_TARGET = "Không tìm thấy dòng quyết toán tương ứng"
REASON_MANY_TARGETS = "Khớp nhiều dòng quyết toán, chưa rõ SQT nào"
REASON_MANY_BILLS = "Container có nhiều Bill, chưa rõ Bill nào đúng"

_ACCENTED_VIETNAMESE = re.compile(
    r"[ÀÁẢÃẠĂẰẮẲẴẶÂẦẤẨẪẬÈÉẺẼẸÊỀẾỂỄỆÌÍỈĨỊ"
    r"ÒÓỎÕỌÔỒỐỔỖỘƠỜỚỞỠỢÙÚỦŨỤƯỪỨỬỮỰỲÝỶỸỴĐ]",
    re.IGNORECASE,
)

DEFAULT_CARGO_MAPPINGS = {
    "VOI": "Vôi",
    "VOI ROI": "Vôi rời",
    "VOI BOT": "Vôi bột",
    "VOI BOT NONG NGHIEP": "Vôi bột nông nghiệp",
    "BOT NN BAO 25 KG": "Bột NN bao 25 kg",
}


class DailyImportError(RuntimeError):
    """Lỗi nghiệp vụ có nội dung phù hợp để hiển thị cho người dùng."""


@dataclass
class ExtractedRow:
    staged_id: int = 0
    json_index: int = -1
    source_row: int = 0
    row_key: str = ""
    source_name: str = ""
    md5: str = ""
    doc_type: str = ""
    source_status: str = ""
    close_date: Optional[str] = None
    container: str = ""
    tons: Optional[float] = None
    cargo: str = ""
    cargo_recognized: bool = True
    place: str = ""
    recipient: str = ""
    truck_no: str = ""
    vessel: str = ""
    sail_date: Optional[str] = None
    carrier: str = ""
    material_price: Optional[float] = None
    invoice_no: str = ""
    bill_no: str = ""
    seal: str = ""
    unit_price: Optional[float] = None
    amount: Optional[float] = None
    vat: Optional[float] = None
    total: Optional[float] = None
    other: str = ""
    confidence: str = ""
    warning: str = ""
    md5_is_synthetic: bool = False

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass
class UnmatchedRow:
    """Một dòng bóc tách không ghép được với dữ liệu quyết toán.

    ``json_index`` là vị trí của dòng trong mảng ``du_lieu_boc_tach`` của file
    JSON tạm, để Bước 2 tô sáng đúng dòng và để xóa đúng dòng khi người dùng
    chọn bỏ các dòng lỗi.
    """

    json_index: int
    staged_id: int
    doc_type: str
    container: str
    match_date: Optional[str]
    reference: str
    reason: str

    @property
    def doc_label(self) -> str:
        return DOC_LABELS.get(self.doc_type, self.doc_type or "Chứng từ")


@dataclass
class FieldConflict:
    conflict_id: str
    target_row: int
    field_name: str
    current_value: Any
    new_value: Any


@dataclass
class InfoChange:
    action: str
    target_row: Optional[int]
    sqt: int
    values: Dict[str, Any]
    staged_ids: List[int]
    document_md5s: List[str]
    check_status: str = "OK"
    conflicts: List[FieldConflict] = field(default_factory=list)


@dataclass
class ExpenseChange:
    action: str
    target_row: Optional[int]
    values: Dict[str, Any]
    staged_ids: List[int]
    document_md5s: List[str]
    check_status: str = "OK"


@dataclass
class ImportAnalysis:
    output_path: str
    daily_path: str
    processed_file_id: Optional[int]
    extracted_rows: int = 0
    duplicate_documents: int = 0
    info_changes: List[InfoChange] = field(default_factory=list)
    expense_changes: List[ExpenseChange] = field(default_factory=list)
    unmatched_rows: List[UnmatchedRow] = field(default_factory=list)
    conflicts: List[FieldConflict] = field(default_factory=list)
    pending_count: int = 0
    pending_states: Dict[int, Tuple[str, str]] = field(default_factory=dict)
    completed_staged_ids: List[int] = field(default_factory=list)

    @property
    def new_info_count(self) -> int:
        return sum(change.action == "CREATE" for change in self.info_changes)

    @property
    def updated_info_count(self) -> int:
        return sum(change.action == "UPDATE" for change in self.info_changes)

    @property
    def has_changes(self) -> bool:
        return bool(self.info_changes or self.expense_changes)


@dataclass
class ImportSummary:
    new_info: int = 0
    updated_info: int = 0
    new_expenses: int = 0
    updated_expenses: int = 0
    duplicate_documents: int = 0
    pending_count: int = 0
    backup_path: str = ""
    status: str = "DAILY_IMPORTED"

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass
class _TargetInfoRow:
    row_number: int
    sqt: int
    close_date: Optional[str]
    container: str
    cargo: str
    values: Dict[str, Any]


@dataclass
class _WorkbookSnapshot:
    info_rows: List[_TargetInfoRow]
    expense_rows: List[Tuple[int, Dict[str, Any]]]
    target_md5: set[str]
    max_sqt: int


def _strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value or "")
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


def _key_text(value: Any) -> str:
    text = _strip_accents(str(value or "")).upper().replace("Đ", "D")
    return re.sub(r"\s+", " ", text).strip()


def normalize_container(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def is_valid_container(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Z]{4}\d{7}", value or ""))


def normalize_md5(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").lower())


def parse_date(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def excel_date(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        parsed = date.fromisoformat(value)
    except ValueError:
        return None
    return datetime(parsed.year, parsed.month, parsed.day)


def parse_number(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(" ", "")
    text = re.sub(r"[^0-9,.-]", "", text)
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        text = "".join(parts) if len(parts[-1]) == 3 else text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def normalize_cargo(
    value: Any, mappings: Optional[Dict[str, str]] = None
) -> Tuple[str, bool]:
    original = re.sub(r"\s+", " ", str(value or "")).strip()
    if not original:
        return "", False
    mapping = dict(DEFAULT_CARGO_MAPPINGS)
    mapping.update(mappings or {})
    key = _key_text(original)
    normalized_mapping = {_key_text(k): v for k, v in mapping.items()}
    if key in normalized_mapping:
        return normalized_mapping[key], True
    if _ACCENTED_VIETNAMESE.search(original):
        return original, True
    replacements = [
        (r"^VOI BOT NONG NGHIEP\b", "Vôi bột nông nghiệp"),
        (r"^VOI CUC\b", "Vôi cục"),
        (r"^VOI CANXI\b", "Vôi canxi"),
        (r"^VOI ROI\b", "Vôi rời"),
        (r"^VOI BOT\b", "Vôi bột"),
        (r"^VOI\b", "Vôi"),
    ]
    for pattern, replacement in replacements:
        if re.search(pattern, key):
            tail = re.sub(pattern, "", key).strip()
            tail = tail.replace(" XA ROI", " xá rời").replace(" CO CHAN", " có chặn")
            return (replacement + (" " + tail.lower() if tail else "")).strip(), not bool(tail)
    return original, False


def classify_doc_type(value: Any) -> str:
    """Nhận diện loại chứng từ từ trường ``loai_chung_tu`` của JSON."""
    key = _key_text(value)
    if "PHIEU CAN" in key:
        return DOC_SCALE
    if key == "BILL" or "BILL OF LADING" in key:
        return DOC_BILL
    if "KHOAN CHI" in key or "CHI PHI" in key:
        return DOC_EXPENSE
    return ""


def match_date_of(item: Dict[str, Any]) -> Optional[str]:
    """Ngày dùng để ghép của một dòng JSON bóc tách (dạng ISO).

    Phiếu cân ghép theo ngày đóng; Khoản chi và Bill ưu tiên ngày chạy rồi mới
    tới ngày đóng. Bước 2 hiển thị đúng ngày này để người dùng biết phần mềm
    đang dùng ngày nào khi ghép dữ liệu.
    """
    close_date = parse_date(item.get("ngay_dong"))
    sail_date = parse_date(item.get("ngay_chay"))
    if classify_doc_type(item.get("loai_chung_tu")) == DOC_SCALE:
        return close_date
    return sail_date or close_date


# ---------------------------------------------------------------------------
# File JSON tạm (bộ nhớ tạm của một lô bóc tách)
# ---------------------------------------------------------------------------
def load_extract_payload(path: str) -> Dict[str, Any]:
    """Đọc file JSON bóc tách và kiểm tra cấu trúc tối thiểu."""
    try:
        with open(path, "r", encoding="utf-8-sig") as f:
            payload = json.load(f)
    except json.JSONDecodeError as exc:
        raise DailyImportError(f"File JSON bóc tách không hợp lệ: {exc}") from exc
    except UnicodeDecodeError as exc:
        raise DailyImportError(
            "File JSON bóc tách không đọc được bằng UTF-8. Vui lòng tải lại file."
        ) from exc
    except OSError as exc:
        raise DailyImportError(f"Không đọc được file JSON bóc tách: {exc}") from exc

    if not isinstance(payload, dict):
        raise DailyImportError("File JSON bóc tách phải có object gốc.")
    rows = payload.get("du_lieu_boc_tach")
    if not isinstance(rows, list):
        raise DailyImportError("File JSON thiếu mảng 'du_lieu_boc_tach'.")
    if any(not isinstance(row, dict) for row in rows):
        raise DailyImportError("Mỗi dòng trong 'du_lieu_boc_tach' phải là object JSON.")
    return payload


def save_extract_payload(path: str, payload: Dict[str, Any]) -> None:
    """Ghi lại file JSON tạm sau khi người dùng sửa/xóa dòng."""
    refresh_extract_metadata(payload)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
            f.write("\n")
    except OSError as exc:
        raise DailyImportError(f"Không lưu được file JSON bóc tách: {exc}") from exc


def refresh_extract_metadata(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Tính lại các con số thống kê trong ``metadata`` theo dữ liệu hiện có."""
    rows = [row for row in payload.get("du_lieu_boc_tach") or [] if isinstance(row, dict)]
    metadata = payload.get("metadata")
    if not isinstance(metadata, dict):
        metadata = {}
        payload["metadata"] = metadata

    counts = extract_summary(rows, payload.get("canh_bao"))
    metadata["tong_so_file"] = counts["files"]
    metadata["tong_so_dong_boc_tach"] = counts["total"]
    metadata["so_dong_ok"] = counts["ok"]
    metadata["so_dong_can_kiem_tra"] = counts["need_check"]
    metadata["tong_so_canh_bao"] = counts["warnings"]
    return metadata


def extract_summary(
    rows: Sequence[Dict[str, Any]],
    global_warnings: Any = None,
    error_indexes: Optional[Iterable[int]] = None,
) -> Dict[str, int]:
    """Thống kê nhanh một lô bóc tách để hiển thị ở đầu màn hình Bước 2.

    ``error_indexes`` là các dòng Bước 3 không nhập được; chúng luôn được tính
    vào nhóm “cần kiểm tra” dù trạng thái trong JSON vẫn là OK.
    """
    errors = set(error_indexes or ())
    warnings = len(global_warnings) if isinstance(global_warnings, list) else 0
    counts = {
        "files": len(
            {str(row.get("file_nguon") or "") for row in rows if row.get("file_nguon")}
        ),
        "total": len(rows),
        "ok": 0,
        "need_check": 0,
        "warnings": warnings,
        DOC_SCALE: 0,
        DOC_BILL: 0,
        DOC_EXPENSE: 0,
        "other": 0,
    }
    for index, row in enumerate(rows):
        counts[classify_doc_type(row.get("loai_chung_tu")) or "other"] += 1
        row_warnings = row.get("canh_bao")
        if isinstance(row_warnings, list):
            counts["warnings"] += len(row_warnings)
        elif row_warnings not in (None, ""):
            counts["warnings"] += 1
        if index not in errors and _key_text(row.get("trang_thai")) in ("", "OK"):
            counts["ok"] += 1
        else:
            counts["need_check"] += 1
    return counts


def remove_extract_rows(path: str, indexes: Iterable[int]) -> int:
    """Xóa các dòng theo vị trí khỏi ``du_lieu_boc_tach`` của file JSON tạm.

    Dùng khi người dùng chọn “Hủy các dòng lỗi và nhập các dòng còn lại”: các
    dòng lỗi biến mất hẳn khỏi bộ nhớ tạm thay vì nằm lại hàng chờ.
    Trả về số dòng đã xóa.
    """
    payload = load_extract_payload(path)
    rows = payload.get("du_lieu_boc_tach") or []
    drop = {index for index in indexes if 0 <= index < len(rows)}
    if not drop:
        return 0
    payload["du_lieu_boc_tach"] = [
        row for index, row in enumerate(rows) if index not in drop
    ]
    save_extract_payload(path, payload)
    return len(drop)


def _join_md5(*values: Any) -> str:
    result: List[str] = []
    for value in values:
        for part in re.split(r"[;,\n]+", str(value or "")):
            md5 = normalize_md5(part)
            if md5 and md5 not in result:
                result.append(md5)
    return ";".join(result)


def _md5_cell_value(values: Dict[str, Any]) -> Any:
    for header in MD5_HEADER_ALIASES:
        if values.get(header) not in (None, ""):
            return values.get(header)
    return ""


def _normalize_import_status(value: Any, default: str = "Chưa nhập") -> str:
    key = _key_text(value)
    if key in ("DA NHAP", "HOAN THANH"):
        return "Đã nhập"
    if key == "CHUA NHAP":
        return "Chưa nhập"
    return default


def _stable_row_key(row: ExtractedRow, occurrence: int = 0) -> str:
    identity = {
        "type": row.doc_type,
        "container": row.container,
        "close_date": row.close_date,
        "sail_date": row.sail_date,
        "bill_no": row.bill_no,
        "invoice_no": row.invoice_no,
        "seal": row.seal,
        "cargo": row.cargo,
        "tons": row.tons,
        "amount": row.amount,
        "total": row.total,
        "occurrence": occurrence,
    }
    raw = json.dumps(identity, ensure_ascii=False, sort_keys=True)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


class DailyImportService:
    def __init__(
        self,
        database: Database,
        logger,
        cargo_mappings: Optional[Dict[str, str]] = None,
    ):
        self.database = database
        self.logger = logger
        self.cargo_mappings = cargo_mappings or {}

    # ------------------------------------------------------------------ #
    # Phân tích
    # ------------------------------------------------------------------ #
    def analyze(
        self,
        output_path: str,
        daily_path: str,
        processed_file_id: Optional[int] = None,
    ) -> ImportAnalysis:
        if not os.path.isfile(output_path):
            raise DailyImportError("Không tìm thấy file JSON bóc tách đang dùng.")
        if not os.path.isfile(daily_path):
            raise DailyImportError(
                "Không tìm thấy file theo dõi hàng ngày. Vui lòng kiểm tra Cài đặt."
            )
        if file_utils.is_file_locked(daily_path):
            raise DailyImportError(
                "File theo dõi hàng ngày đang mở. Vui lòng lưu và đóng Excel rồi thử lại."
            )

        parsed = self._read_output(output_path)
        snapshot = self._read_daily_snapshot(daily_path)
        analysis = ImportAnalysis(
            output_path=output_path,
            daily_path=daily_path,
            processed_file_id=processed_file_id,
            extracted_rows=len(parsed),
        )

        # JSON tạm là nguồn dữ liệu duy nhất của lần nhập này: dọn hàng chờ cũ
        # rồi chỉ ghi nhận đúng các dòng đang có trong file.
        self.database.clear_pending_staged_rows()
        occurrences: Dict[Tuple[str, str], int] = {}
        for row in parsed:
            if not row.md5:
                continue
            key = (row.md5, _stable_row_key(row))
            occurrences[key] = occurrences.get(key, 0) + 1
            row.row_key = _stable_row_key(row, occurrences[key] - 1)
            self.database.upsert_source_document(
                row.md5,
                row.doc_type,
                row.source_name,
                processed_file_id,
            )
            row.staged_id = self.database.upsert_staged_row(
                row.md5,
                row.row_key,
                row.source_row,
                row.doc_type,
                row.container,
                row.to_dict(),
            )

        # Một dòng chỉ được coi là "đã nhập" khi CẢ HAI cùng đồng ý:
        #   - database: đúng dòng đó đã được ghi xong (một file chứng từ có thể sinh
        #     nhiều dòng, ví dụ một ảnh khoản chi có 4 container);
        #   - file theo dõi: MD5 của chứng từ vẫn còn trong cột MD5.
        # Nhờ vế thứ hai, người dùng xóa dòng khỏi Excel là nhập lại được ngay mà
        # không phải đụng vào database.
        never_written = {
            int(item["id"])
            for item in self.database.list_staged_rows(include_completed=False)
        }
        active: List[ExtractedRow] = []
        duplicate_md5: set[str] = set()
        for row in parsed:
            if not row.staged_id:
                continue
            written_before = row.staged_id not in never_written
            # Chứng từ không có MD5 thì không soi được trong file theo dõi; đành chỉ
            # dựa vào bộ nhớ của app để khỏi ghi trùng.
            still_in_daily = row.md5_is_synthetic or row.md5 in snapshot.target_md5
            if written_before and still_in_daily:
                if not row.md5_is_synthetic:
                    duplicate_md5.add(row.md5)
                continue
            active.append(row)
        analysis.duplicate_documents = len(duplicate_md5)

        bills_by_container: Dict[str, List[ExtractedRow]] = {}
        scales: List[ExtractedRow] = []
        expenses: List[ExtractedRow] = []
        for row in active:
            if row.doc_type == DOC_BILL:
                bills_by_container.setdefault(row.container, []).append(row)
            elif row.doc_type == DOC_SCALE:
                scales.append(row)
            elif row.doc_type == DOC_EXPENSE:
                expenses.append(row)

        next_sqt = snapshot.max_sqt + 1
        used_bills: set[int] = set()
        resolved_bill_containers: set[str] = set()
        blocked_containers: set[str] = set()
        virtual_rows = list(snapshot.info_rows)

        for scale in scales:
            raw_bills = bills_by_container.get(scale.container, [])
            candidates = self._distinct_bills(raw_bills)
            if len(candidates) > 1:
                # Không tự đoán Bill nào đúng: người dùng sửa lại ở Bước 2. Báo cả
                # Phiếu cân lẫn mọi dòng Bill của container để bỏ là bỏ hết.
                self._mark_unmatched(analysis, scale, REASON_MANY_BILLS)
                for bill in raw_bills:
                    self._mark_unmatched(analysis, bill, REASON_MANY_BILLS)
                blocked_containers.add(scale.container)
                continue

            # Container khớp nhiều dòng quyết toán: Bill của container cũng sẽ khớp
            # nhiều dòng như vậy nên cứ để vòng Bill bên dưới báo lỗi cho nó.
            target_options = self._scale_target_options(virtual_rows, scale)
            if len(target_options) > 1:
                self._mark_unmatched(analysis, scale, REASON_MANY_TARGETS)
                continue

            selected_bill = candidates[0] if candidates else None
            target = target_options[0] if target_options else None
            fallback = target.values if target else {}
            sqt = target.sqt if target else next_sqt
            action = "UPDATE" if target else "CREATE"
            if not target:
                next_sqt += 1
            previous_same_container = any(
                item.container == scale.container
                and item.close_date != scale.close_date
                for item in snapshot.info_rows
            )
            check_status = (
                "Cần kiểm tra"
                if previous_same_container
                or not scale.cargo_recognized
                # Tạo QT mới mà không có Bill (thông tin tàu nhập tay) -> đánh dấu
                # để người dùng soát lại cho an toàn.
                or (not selected_bill and not target)
                or self._source_needs_review(scale, selected_bill)
                else "OK"
            )
            values = self._info_values(scale, selected_bill, sqt, fallback)
            conflicts = self._find_conflicts(target, values) if target else []
            if conflicts:
                check_status = "Cần kiểm tra"
                analysis.conflicts.extend(conflicts)
            staged_ids = [scale.staged_id]
            md5_values = [self._real_md5(scale)]
            if selected_bill:
                staged_ids.append(selected_bill.staged_id)
                md5_values.append(self._real_md5(selected_bill))
                used_bills.add(selected_bill.staged_id)
                resolved_bill_containers.add(scale.container)
                analysis.pending_states[selected_bill.staged_id] = (STATE_READY, "")
            analysis.info_changes.append(
                InfoChange(
                    action=action,
                    target_row=target.row_number if target else None,
                    sqt=sqt,
                    values=values,
                    staged_ids=staged_ids,
                    document_md5s=md5_values,
                    check_status=check_status,
                    conflicts=conflicts,
                )
            )
            analysis.pending_states[scale.staged_id] = (STATE_READY, "")
            analysis.completed_staged_ids.extend(staged_ids)
            if action == "CREATE":
                virtual_rows.append(
                    _TargetInfoRow(
                        row_number=0,
                        sqt=sqt,
                        close_date=scale.close_date,
                        container=scale.container,
                        cargo=scale.cargo,
                        values=values,
                    )
                )

        # Bill có thể bổ sung trực tiếp cho một dòng quyết toán đã tồn tại.
        for container, raw_candidates in bills_by_container.items():
            if container in blocked_containers:
                continue
            remaining = [
                row for row in raw_candidates if row.staged_id not in used_bills
            ]
            if not remaining:
                continue
            if container in resolved_bill_containers:
                # Cùng một Bill (trùng MD5) đã ghép với Phiếu cân ở trên.
                for bill in remaining:
                    analysis.pending_states[bill.staged_id] = (STATE_READY, "")
                    analysis.completed_staged_ids.append(bill.staged_id)
                continue

            candidates = self._distinct_bills(remaining)
            if len(candidates) > 1:
                for bill in remaining:
                    self._mark_unmatched(analysis, bill, REASON_MANY_BILLS)
                continue
            selected_bill = candidates[0]

            target_rows = [row for row in virtual_rows if row.container == container]
            if len(target_rows) > 1:
                self._mark_unmatched(analysis, selected_bill, REASON_MANY_TARGETS)
                continue
            target = target_rows[0] if target_rows else None
            action = "UPDATE" if target else "CREATE"
            sqt = target.sqt if target else next_sqt
            if not target:
                next_sqt += 1
            fallback = target.values if target else {}
            values = self._bill_info_values(selected_bill, sqt, fallback)
            conflicts = self._find_conflicts(target, values) if target else []
            analysis.conflicts.extend(conflicts)
            analysis.info_changes.append(
                InfoChange(
                    action=action,
                    target_row=target.row_number if target else None,
                    sqt=sqt,
                    values=values,
                    staged_ids=[selected_bill.staged_id],
                    document_md5s=[self._real_md5(selected_bill)],
                    check_status="OK",
                    conflicts=conflicts,
                )
            )
            if action == "CREATE":
                virtual_rows.append(
                    _TargetInfoRow(
                        row_number=0,
                        sqt=sqt,
                        close_date=selected_bill.close_date,
                        container=selected_bill.container,
                        cargo=selected_bill.cargo,
                        values=values,
                    )
                )
            used_bills.add(selected_bill.staged_id)
            analysis.pending_states[selected_bill.staged_id] = (STATE_READY, "")
            analysis.completed_staged_ids.append(selected_bill.staged_id)

        self._analyze_expenses(expenses, virtual_rows, snapshot, analysis)

        for row_id, (state, note) in analysis.pending_states.items():
            self.database.update_staged_row(row_id, state=state, note=note)
        # Dòng không ghép được không nằm lại hàng chờ: người dùng sẽ sửa ở Bước 2
        # hoặc bỏ hẳn ngay trong lần nhập này.
        stale = [
            row.staged_id
            for row in active
            if row.staged_id not in analysis.pending_states
        ]
        if stale:
            self.database.delete_staged_rows(stale)
        for md5 in {row.md5 for row in parsed if row.md5}:
            self.database.refresh_document_status(md5)
        analysis.pending_count = len(analysis.unmatched_rows)
        return analysis

    # ------------------------------------------------------------------ #
    # Ghi workbook
    # ------------------------------------------------------------------ #
    def commit(
        self,
        analysis: ImportAnalysis,
        conflict_decisions: Optional[Dict[str, bool]] = None,
    ) -> ImportSummary:
        conflict_decisions = conflict_decisions or {}
        daily_path = analysis.daily_path
        if file_utils.is_file_locked(daily_path):
            raise DailyImportError(
                "File theo dõi hàng ngày đang mở. Vui lòng đóng Excel rồi thử lại."
            )
        run_id = self.database.create_import_run(
            analysis.processed_file_id, analysis.output_path, daily_path
        )
        backup_path = ""
        temp_path = ""
        try:
            wb = load_workbook(daily_path)
            self._upgrade_workbook(wb)
            info_ws = wb[INFO_SHEET]
            expense_ws = wb[EXPENSE_SHEET]
            info_map = self._header_map(info_ws)
            expense_map = self._header_map(expense_ws)

            for change in analysis.info_changes:
                row_number = (
                    change.target_row
                    if change.action == "UPDATE" and change.target_row
                    else info_ws.max_row + 1
                )
                if change.action == "CREATE":
                    self._copy_row_style(info_ws, max(2, info_ws.max_row), row_number)
                conflict_by_field = {
                    conflict.field_name: conflict for conflict in change.conflicts
                }
                for name, value in change.values.items():
                    if name not in info_map:
                        continue
                    cell = info_ws.cell(row_number, info_map[name])
                    conflict = conflict_by_field.get(name)
                    if conflict:
                        if conflict_decisions.get(conflict.conflict_id, False):
                            cell.value = value
                    elif change.action == "CREATE" or cell.value in (None, "") or name in (
                        "MD5",
                        "Ngày cập nhật",
                    ):
                        cell.value = value

            for change in analysis.expense_changes:
                row_number = (
                    change.target_row
                    if change.action == "UPDATE" and change.target_row
                    else expense_ws.max_row + 1
                )
                if change.action == "CREATE":
                    self._copy_row_style(expense_ws, max(2, expense_ws.max_row), row_number)
                for name, value in change.values.items():
                    if name not in expense_map:
                        continue
                    cell = expense_ws.cell(row_number, expense_map[name])
                    if change.action == "CREATE" or cell.value in (None, "") or name in (
                        "MD5",
                        "Ngày cập nhật",
                    ):
                        cell.value = value

            self._format_workbook(wb)
            suffix = Path(daily_path).suffix or ".xlsx"
            temp_path = str(Path(daily_path).with_name(f".__daily_import_{os.getpid()}{suffix}"))
            wb.save(temp_path)
            wb.close()
            self._validate_saved_workbook(temp_path)
            os.replace(temp_path, daily_path)

            completed_ids = sorted(set(analysis.completed_staged_ids))
            affected_md5: set[str] = set()
            for row_id in completed_ids:
                item = self.database.get_staged_row(row_id)
                if not item:
                    continue
                affected_md5.add(item.get("document_md5") or "")
                self.database.update_staged_row(row_id, state=STATE_COMPLETED)
            for md5 in affected_md5:
                if md5:
                    self.database.refresh_document_status(md5)

            pending = self.database.count_pending_rows()
            status = "DAILY_IMPORT_PARTIAL" if pending else "DAILY_IMPORTED"
            summary = ImportSummary(
                new_info=analysis.new_info_count,
                updated_info=analysis.updated_info_count,
                new_expenses=sum(c.action == "CREATE" for c in analysis.expense_changes),
                updated_expenses=sum(c.action == "UPDATE" for c in analysis.expense_changes),
                duplicate_documents=analysis.duplicate_documents,
                pending_count=pending,
                backup_path=backup_path,
                status=status,
            )
            self.database.finish_import_run(
                run_id, "COMPLETED", backup_path=backup_path, summary=summary.to_dict()
            )
            if analysis.processed_file_id is not None:
                self.database.update_status(
                    analysis.processed_file_id,
                    status,
                    note=self._summary_note(summary),
                )
            return summary
        except Exception as exc:
            # Ghi theo cơ chế file tạm + os.replace (nguyên tử) nên file theo dõi
            # gốc không bị hỏng dở; chỉ cần dọn file tạm nếu lỗi.
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except OSError:
                    pass
            self.database.finish_import_run(
                run_id,
                "FAILED",
                backup_path=backup_path or None,
                error_message=str(exc),
            )
            if analysis.processed_file_id is not None:
                self.database.update_status(
                    analysis.processed_file_id,
                    "DAILY_IMPORT_ERROR",
                    note=f"Lỗi nhập file theo dõi: {exc}",
                )
            if isinstance(exc, DailyImportError):
                raise
            raise DailyImportError(f"Không cập nhật được file theo dõi: {exc}") from exc

    # ------------------------------------------------------------------ #
    # Đọc dữ liệu
    # ------------------------------------------------------------------ #
    def _read_output(self, path: str) -> List[ExtractedRow]:
        if Path(path).suffix.lower() != ".json":
            raise DailyImportError("File bóc tách phải là định dạng JSON (.json).")

        payload = load_extract_payload(path)
        rows = payload.get("du_lieu_boc_tach") or []

        result: List[ExtractedRow] = []
        for index, item in enumerate(rows, start=1):
            if not any(value not in (None, "", [], {}) for value in item.values()):
                continue

            doc_type = classify_doc_type(item.get("loai_chung_tu"))
            if not doc_type:
                continue

            cargo, recognized = normalize_cargo(
                item.get("loai_hang"),
                self.cargo_mappings,
            )
            close_date = parse_date(item.get("ngay_dong"))
            sail_date = parse_date(item.get("ngay_chay"))
            source_row_number = parse_number(item.get("stt_hien_thi"))
            row = ExtractedRow(
                json_index=index - 1,
                source_row=int(source_row_number or index),
                source_name=str(item.get("file_nguon") or ""),
                md5=normalize_md5(item.get("ma_md5_file")),
                doc_type=doc_type,
                source_status=str(item.get("trang_thai") or ""),
                close_date=close_date,
                container=normalize_container(item.get("so_container")),
                tons=parse_number(item.get("so_tan")),
                cargo=cargo,
                cargo_recognized=recognized,
                place=str(item.get("noi_dong") or "").strip(),
                recipient=str(item.get("nguoi_nhan") or "").strip(),
                truck_no=str(item.get("bien_so_xe") or "").strip(),
                vessel=str(item.get("ten_tau") or "").strip(),
                sail_date=sail_date,
                carrier=str(item.get("vt_bien") or "").strip(),
                material_price=parse_number(item.get("gia_vat_lieu")),
                invoice_no=str(item.get("so_hd") or "").strip(),
                bill_no=str(item.get("so_bill") or "").strip(),
                seal=str(item.get("so_chi_seal") or "").strip(),
                unit_price=parse_number(item.get("don_gia")),
                amount=parse_number(item.get("thanh_tien")),
                vat=parse_number(item.get("vat")),
                total=parse_number(item.get("tong_tien")),
                other=self._json_cell_text(item.get("truong_khac")),
                confidence=str(item.get("do_tin_cay") or "").strip(),
                warning=self._json_warning_text(item.get("canh_bao")),
            )
            if doc_type == DOC_EXPENSE:
                row.close_date = sail_date or close_date
            if not row.md5:
                # Không có MD5 thì vẫn theo dõi được trong danh sách chờ, nhưng
                # dùng khóa cục bộ để không làm mất dòng.
                row.md5 = "missing-" + hashlib.sha1(
                    f"{path}|{index}|{row.container}".encode("utf-8")
                ).hexdigest()
                row.md5_is_synthetic = True
            result.append(row)
        return result

    @staticmethod
    def _json_cell_text(value: Any) -> str:
        if value in (None, ""):
            return ""
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        return str(value).strip()

    @classmethod
    def _json_warning_text(cls, value: Any) -> str:
        if value in (None, ""):
            return ""
        if isinstance(value, list):
            return "; ".join(
                cls._json_cell_text(item) for item in value if item not in (None, "")
            )
        return cls._json_cell_text(value)

    def _read_daily_snapshot(self, path: str) -> _WorkbookSnapshot:
        try:
            wb = load_workbook(path, read_only=True, data_only=False)
        except Exception as exc:
            raise DailyImportError(f"Không đọc được file theo dõi hàng ngày: {exc}") from exc
        try:
            for sheet_name in (INFO_SHEET, EXPENSE_SHEET):
                if sheet_name not in wb.sheetnames:
                    raise DailyImportError(
                        f"File theo dõi thiếu sheet '{sheet_name}'."
                    )
            info_ws = wb[INFO_SHEET]
            expense_ws = wb[EXPENSE_SHEET]
            info_map = self._validate_daily_headers(info_ws, INFO_HEADERS)
            expense_map = self._validate_daily_headers(expense_ws, EXPENSE_HEADERS)
            info_rows: List[_TargetInfoRow] = []
            target_md5: set[str] = set()
            max_sqt = 0
            for row_number in range(2, info_ws.max_row + 1):
                values = {
                    name: info_ws.cell(row_number, col).value
                    for name, col in info_map.items()
                }
                if not any(value not in (None, "") for value in values.values()):
                    continue
                sqt_value = parse_number(values.get("SQT PM"))
                sqt = int(sqt_value or 0)
                max_sqt = max(max_sqt, sqt)
                container = normalize_container(values.get("Số Container"))
                cargo, _ = normalize_cargo(values.get("Loại hàng"), self.cargo_mappings)
                info_rows.append(
                    _TargetInfoRow(
                        row_number=row_number,
                        sqt=sqt,
                        close_date=parse_date(values.get("Ngày Đóng")),
                        container=container,
                        cargo=cargo,
                        values=values,
                    )
                )
                target_md5.update(
                    normalize_md5(part)
                    for part in re.split(r"[;,\n]+", str(_md5_cell_value(values) or ""))
                    if normalize_md5(part)
                )
            expense_rows: List[Tuple[int, Dict[str, Any]]] = []
            for row_number in range(2, expense_ws.max_row + 1):
                values = {
                    name: expense_ws.cell(row_number, col).value
                    for name, col in expense_map.items()
                }
                if not any(value not in (None, "") for value in values.values()):
                    continue
                expense_rows.append((row_number, values))
                target_md5.update(
                    normalize_md5(part)
                    for part in re.split(r"[;,\n]+", str(_md5_cell_value(values) or ""))
                    if normalize_md5(part)
                )
            return _WorkbookSnapshot(info_rows, expense_rows, target_md5, max_sqt)
        finally:
            wb.close()

    # ------------------------------------------------------------------ #
    # Matching helpers
    # ------------------------------------------------------------------ #
    @staticmethod
    def _target_group(
        rows: Sequence[_TargetInfoRow], close_date: Optional[str], container: str
    ) -> List[_TargetInfoRow]:
        return [
            row
            for row in rows
            if row.close_date == close_date and row.container == container
        ]

    @staticmethod
    def _real_md5(row: ExtractedRow) -> str:
        return "" if row.md5_is_synthetic or row.md5.startswith("missing-") else row.md5

    def _scale_target_options(
        self, rows: Sequence[_TargetInfoRow], scale: ExtractedRow
    ) -> List[_TargetInfoRow]:
        group = self._target_group(rows, scale.close_date, scale.container)
        if not group:
            return []
        if scale.cargo and any(row.cargo for row in group):
            return [
                row for row in group if _key_text(row.cargo) == _key_text(scale.cargo)
            ]
        return group

    @staticmethod
    def _distinct_bills(rows: Iterable[ExtractedRow]) -> List[ExtractedRow]:
        result: List[ExtractedRow] = []
        seen = set()
        for row in rows:
            if row.md5 not in seen:
                seen.add(row.md5)
                result.append(row)
        return result

    @staticmethod
    def _match_date(row: ExtractedRow) -> Optional[str]:
        """Ngày dùng để ghép của một dòng đã bóc tách (Bill ưu tiên ngày chạy)."""
        if row.doc_type == DOC_BILL:
            return row.sail_date or row.close_date
        return row.close_date

    @staticmethod
    def _reference_text(row: ExtractedRow) -> str:
        """Số HĐ/Bill/Seal để người dùng nhận ra dòng lỗi trong file gốc."""
        if row.doc_type == DOC_EXPENSE and row.invoice_no:
            return f"HĐ {row.invoice_no}"
        if row.doc_type == DOC_BILL and row.bill_no:
            return f"Bill {row.bill_no}"
        if row.doc_type == DOC_SCALE and row.seal:
            return f"Seal {row.seal}"
        return row.invoice_no or row.bill_no or row.seal or ""

    def _mark_unmatched(
        self, analysis: ImportAnalysis, row: ExtractedRow, reason: str
    ) -> None:
        analysis.unmatched_rows.append(
            UnmatchedRow(
                json_index=row.json_index,
                staged_id=row.staged_id,
                doc_type=row.doc_type,
                container=row.container,
                match_date=self._match_date(row),
                reference=self._reference_text(row),
                reason=reason,
            )
        )

    @staticmethod
    def _source_needs_review(*rows: Optional[ExtractedRow]) -> bool:
        for row in rows:
            if row is None:
                continue
            if row.warning:
                return True
            if _key_text(row.source_status) not in ("", "OK"):
                return True
            if _key_text(row.confidence) == "THAP":
                return True
        return False

    def _info_values(
        self,
        scale: ExtractedRow,
        bill: Optional[ExtractedRow],
        sqt: int,
        fallback: Dict[str, Any],
    ) -> Dict[str, Any]:
        vessel = (
            (bill.vessel if bill else "")
            or str(fallback.get("Tên tàu") or "")
            or (scale.vessel or "")
        )
        sail_date = (
            (bill.sail_date if bill else None)
            or parse_date(fallback.get("Ngày chạy"))
            or scale.sail_date
        )
        carrier = (
            (bill.carrier if bill else "")
            or str(fallback.get("VT biển") or "")
            or (scale.carrier or "")
        )
        seal = scale.seal or (bill.seal if bill else "") or fallback.get("Số chì/Seal")
        return {
            "SQT PM": sqt,
            "Ngày Đóng": excel_date(scale.close_date),
            "Số Container": scale.container,
            "Biển số xe": scale.truck_no or fallback.get("Biển số xe"),
            "Số chì/Seal": seal,
            "Số tấn": scale.tons,
            "Loại hàng": scale.cargo,
            "Nơi đóng": scale.place or fallback.get("Nơi đóng"),
            "Tên tàu": vessel,
            "Ngày chạy": excel_date(sail_date),
            "Người nhận": scale.recipient or fallback.get("Người nhận"),
            "VT biển": carrier,
            "MD5": _join_md5(
                _md5_cell_value(fallback),
                self._real_md5(scale),
                self._real_md5(bill) if bill else "",
            ),
            "Trạng thái nhập": _normalize_import_status(
                fallback.get("Trạng thái nhập"), "Chưa nhập"
            ),
            "Ngày cập nhật": datetime.now(),
        }

    def _bill_info_values(
        self, bill: ExtractedRow, sqt: int, fallback: Dict[str, Any]
    ) -> Dict[str, Any]:
        return {
            "SQT PM": sqt,
            "Ngày Đóng": excel_date(bill.close_date) or fallback.get("Ngày Đóng"),
            "Số Container": bill.container or fallback.get("Số Container"),
            "Biển số xe": bill.truck_no or fallback.get("Biển số xe"),
            "Số chì/Seal": bill.seal or fallback.get("Số chì/Seal"),
            "Số tấn": bill.tons if bill.tons is not None else fallback.get("Số tấn"),
            "Loại hàng": bill.cargo or fallback.get("Loại hàng"),
            "Nơi đóng": bill.place or fallback.get("Nơi đóng"),
            "Tên tàu": bill.vessel or fallback.get("Tên tàu"),
            "Ngày chạy": excel_date(bill.sail_date) or fallback.get("Ngày chạy"),
            "Người nhận": bill.recipient or fallback.get("Người nhận"),
            "VT biển": bill.carrier or fallback.get("VT biển"),
            "MD5": _join_md5(_md5_cell_value(fallback), self._real_md5(bill)),
            "Trạng thái nhập": _normalize_import_status(
                fallback.get("Trạng thái nhập"), "Chưa nhập"
            ),
            "Ngày cập nhật": datetime.now(),
        }

    @staticmethod
    def _find_conflicts(
        target: Optional[_TargetInfoRow], values: Dict[str, Any]
    ) -> List[FieldConflict]:
        if target is None:
            return []
        ignored = {
            "MD5",
            "Trạng thái nhập",
            "Ngày nhập cuối",
            "Ngày cập nhật",
        }
        conflicts = []
        for name, new_value in values.items():
            current = target.values.get(name)
            if name in ignored or current in (None, "") or new_value in (None, ""):
                continue
            current_compare = parse_date(current) or _key_text(current)
            new_compare = parse_date(new_value) or _key_text(new_value)
            if current_compare != new_compare:
                raw_id = f"{target.row_number}|{name}|{current}|{new_value}"
                conflicts.append(
                    FieldConflict(
                        conflict_id=hashlib.sha1(raw_id.encode("utf-8")).hexdigest(),
                        target_row=target.row_number,
                        field_name=name,
                        current_value=current,
                        new_value=new_value,
                    )
                )
        return conflicts

    def _analyze_expenses(
        self,
        expenses: Sequence[ExtractedRow],
        virtual_rows: Sequence[_TargetInfoRow],
        snapshot: _WorkbookSnapshot,
        analysis: ImportAnalysis,
    ) -> None:
        for expense in expenses:
            matches = [
                row
                for row in virtual_rows
                if row.close_date == expense.close_date
                and row.container == expense.container
            ]
            sqt_values = {row.sqt for row in matches if row.sqt}
            if len(sqt_values) != 1:
                # Khoản chi không có đúng một SQT thì không ghi vào sheet Khoan_Chi:
                # ghi với SQT rỗng sẽ tạo dòng mồ côi không đối chiếu được.
                self._mark_unmatched(
                    analysis,
                    expense,
                    REASON_NO_TARGET if not sqt_values else REASON_MANY_TARGETS,
                )
                continue
            sqt = next(iter(sqt_values))
            check_status = (
                "Cần kiểm tra" if self._source_needs_review(expense) else "OK"
            )
            if (
                expense.total is not None
                and expense.amount is not None
                and expense.vat is not None
                and abs(float(expense.total) - float(expense.amount) - float(expense.vat)) > 1
            ):
                check_status = "Cần kiểm tra"
            existing = next(
                (
                    (row_number, values)
                    for row_number, values in snapshot.expense_rows
                    if parse_date(values.get("Ngày tháng")) == expense.close_date
                    and normalize_container(values.get("Số Container")) == expense.container
                ),
                None,
            )
            values = {
                "SQT PM": sqt,
                "Ngày tháng": excel_date(expense.close_date),
                "Số Container": expense.container,
                "Số HĐ": expense.invoice_no,
                "Giá vật liệu": expense.material_price,
                "Đơn giá": expense.unit_price,
                "Thành tiền": expense.amount,
                "VAT": expense.vat,
                "Tổng tiền": expense.total,
                "MD5": _join_md5(
                    _md5_cell_value(existing[1]) if existing else "",
                    self._real_md5(expense),
                ),
                "Trạng thái nhập": _normalize_import_status(
                    existing[1].get("Trạng thái nhập") if existing else None,
                    "Chưa nhập",
                ),
                "Ngày cập nhật": datetime.now(),
            }
            analysis.expense_changes.append(
                ExpenseChange(
                    action="UPDATE" if existing else "CREATE",
                    target_row=existing[0] if existing else None,
                    values=values,
                    staged_ids=[expense.staged_id],
                    document_md5s=[self._real_md5(expense)],
                    check_status=check_status,
                )
            )
            analysis.pending_states[expense.staged_id] = (STATE_READY, "")
            analysis.completed_staged_ids.append(expense.staged_id)

    # ------------------------------------------------------------------ #
    # Workbook helpers
    # ------------------------------------------------------------------ #
    @staticmethod
    def _header_map(ws) -> Dict[str, int]:
        mapping = {
            str(ws.cell(1, col).value): col
            for col in range(1, ws.max_column + 1)
            if ws.cell(1, col).value not in (None, "")
        }
        if "MD5" not in mapping and "Mã MD5" in mapping:
            mapping["MD5"] = mapping["Mã MD5"]
        if "Mã MD5" not in mapping and "MD5" in mapping:
            mapping["Mã MD5"] = mapping["MD5"]
        return mapping

    def _validate_daily_headers(self, ws, expected: Sequence[str]) -> Dict[str, int]:
        mapping = self._header_map(ws)
        auto_addable = {"Biển số xe", "Ngày nhập cuối", "Ngày cập nhật", "MD5"}
        missing = [
            name
            for name in expected
            if name not in mapping and name not in auto_addable
        ]
        if missing:
            raise DailyImportError(
                f"Sheet '{ws.title}' thiếu các cột: " + ", ".join(missing)
            )
        return mapping

    def _upgrade_workbook(self, wb) -> None:
        for sheet_name, expected in (
            (INFO_SHEET, INFO_HEADERS),
            (EXPENSE_SHEET, EXPENSE_HEADERS),
        ):
            if sheet_name not in wb.sheetnames:
                raise DailyImportError(f"File theo dõi thiếu sheet '{sheet_name}'.")
            ws = wb[sheet_name]
            self._validate_daily_headers(ws, expected)
            if sheet_name == INFO_SHEET:
                self._ensure_column(ws, "Biển số xe", before_header="Số chì/Seal")
            self._ensure_column(ws, "Trạng thái nhập")
            self._ensure_column(ws, "Ngày cập nhật", before_header="MD5")
            if sheet_name == INFO_SHEET:
                self._ensure_column(ws, "Ngày nhập cuối", before_header="Ngày cập nhật")
            if "MD5" not in self._header_map(ws):
                self._ensure_column(ws, "MD5")
            mapping = self._header_map(ws)
            missing_after = [name for name in expected if name not in mapping]
            if missing_after:
                raise DailyImportError(
                    f"Không nâng cấp được sheet '{sheet_name}': "
                    + ", ".join(missing_after)
                )

    def _ensure_column(
        self,
        ws,
        header: str,
        *,
        before_header: Optional[str] = None,
    ) -> None:
        mapping = self._header_map(ws)
        if header in mapping:
            return
        insert_col = ws.max_column + 1
        if before_header and before_header in mapping:
            insert_col = mapping[before_header]
            ws.insert_cols(insert_col, 1)
        ws.cell(1, insert_col).value = header
        style_source_col = min(insert_col + 1, ws.max_column)
        if style_source_col != insert_col:
            ws.cell(1, insert_col)._style = copy(ws.cell(1, style_source_col)._style)
        for row_number in range(2, ws.max_row + 1):
            if style_source_col != insert_col:
                ws.cell(row_number, insert_col)._style = copy(
                    ws.cell(row_number, style_source_col)._style
                )

    @staticmethod
    def _copy_row_style(ws, source_row: int, target_row: int) -> None:
        if source_row < 2 or source_row > ws.max_row:
            return
        for col in range(1, ws.max_column + 1):
            source = ws.cell(source_row, col)
            target = ws.cell(target_row, col)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format
            target.alignment = copy(source.alignment)

    def _format_workbook(self, wb) -> None:
        for sheet_name in (INFO_SHEET, EXPENSE_SHEET):
            ws = wb[sheet_name]
            mapping = self._header_map(ws)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(1, ws.max_row)}"
            # Form mới chỉ còn hai trạng thái nhập phục vụ PAD.
            ws.data_validations.dataValidation = []
            import_col = get_column_letter(mapping["Trạng thái nhập"])
            import_validation = DataValidation(
                type="list", formula1='"Chưa nhập,Đã nhập"', allow_blank=False
            )
            ws.add_data_validation(import_validation)
            import_validation.add(f"{import_col}2:{import_col}10000")
            status_col = mapping["Trạng thái nhập"]
            for row_number in range(2, ws.max_row + 1):
                if any(
                    ws.cell(row_number, col).value not in (None, "")
                    for col in range(1, ws.max_column + 1)
                    if col != status_col
                ):
                    ws.cell(row_number, status_col).value = _normalize_import_status(
                        ws.cell(row_number, status_col).value,
                        "Chưa nhập",
                    )
            for header in ("Ngày Đóng", "Ngày chạy", "Ngày Giao", "Ngày tháng"):
                if header in mapping:
                    for row_number in range(2, ws.max_row + 1):
                        ws.cell(row_number, mapping[header]).number_format = "dd/mm/yyyy"
            for header in ("Ngày nhập cuối", "Ngày cập nhật"):
                if header not in mapping:
                    continue
                for row_number in range(2, ws.max_row + 1):
                    ws.cell(row_number, mapping[header]).number_format = (
                        "dd/mm/yyyy hh:mm:ss"
                    )
            for header in ("Giá vật liệu", "Đơn giá", "Thành tiền", "VAT", "Tổng tiền"):
                if header in mapping:
                    for row_number in range(2, ws.max_row + 1):
                        ws.cell(row_number, mapping[header]).number_format = "#,##0"
            if "Số tấn" in mapping:
                for row_number in range(2, ws.max_row + 1):
                    ws.cell(row_number, mapping["Số tấn"]).number_format = "0.00"
            # Cập nhật bảng Excel nếu workbook cũ có Table.
            for table in ws.tables.values():
                table.ref = f"A1:{get_column_letter(ws.max_column)}{max(2, ws.max_row)}"

    def _validate_saved_workbook(self, path: str) -> None:
        wb = load_workbook(path, read_only=True, data_only=False)
        try:
            for sheet_name, expected in (
                (INFO_SHEET, INFO_HEADERS),
                (EXPENSE_SHEET, EXPENSE_HEADERS),
            ):
                if sheet_name not in wb.sheetnames:
                    raise DailyImportError(f"File tạm thiếu sheet '{sheet_name}'.")
                mapping = self._header_map(wb[sheet_name])
                missing = [name for name in expected if name not in mapping]
                if missing:
                    raise DailyImportError(
                        f"File tạm thiếu cột ở sheet '{sheet_name}': "
                        + ", ".join(missing)
                    )
        finally:
            wb.close()

    @staticmethod
    def _summary_note(summary: ImportSummary) -> str:
        return (
            f"Đã nhập file theo dõi: {summary.new_info} dòng quyết toán mới, "
            f"{summary.updated_info} dòng cập nhật, {summary.new_expenses} khoản chi mới; "
            f"còn {summary.pending_count} dữ liệu chờ."
        )
