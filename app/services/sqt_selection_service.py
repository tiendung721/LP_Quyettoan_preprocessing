"""Read SQT choices from the daily workbook and write PAD selection JSON."""

from __future__ import annotations

import json
import math
import os
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, List

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .. import file_utils
from ..daily_import import INFO_SHEET, is_month_info_sheet

SQT_COLUMN = "SQT PM"
SELECTION_OPERATION = "NHAP_THONG_TIN"
EXPENSE_SELECTION_OPERATION = "NHAP_KHOAN_CHI"


class SqtSelectionError(RuntimeError):
    """Business error safe to show directly to the user."""


class DailyFileNotFoundError(SqtSelectionError):
    pass


class DailyFileLockedError(SqtSelectionError):
    pass


class MissingSheetError(SqtSelectionError):
    pass


class MissingColumnError(SqtSelectionError):
    pass


class EmptySqtListError(SqtSelectionError):
    pass


class SelectionJsonWriteError(SqtSelectionError):
    pass


@dataclass(frozen=True)
class SqtSelectionItem:
    value: str
    row_count: int
    sheet_name: str = ""
    row_numbers: tuple[int, ...] = ()

    @property
    def display_text(self) -> str:
        if self.sheet_name:
            return f"{self.value} — {self.row_count} dòng — {self.sheet_name}"
        return f"{self.value} — {self.row_count} dòng"

    def to_selection_payload(self) -> dict[str, Any]:
        payload: dict[str, Any] = {"sqt": self.value}
        if self.sheet_name:
            payload["sheet_name"] = self.sheet_name
        if self.row_numbers:
            payload["row_numbers"] = list(self.row_numbers)
        return payload


def normalize_sqt_value(value: Any) -> str:
    """Return the stable string sent to PAD for an Excel SQT cell value."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return format(value, "g").strip()

    text = str(value).strip()
    if not text:
        return ""
    if text.endswith(".0"):
        head = text[:-2]
        if head and head.lstrip("+-").isdigit():
            return str(int(head))
    return text


def _header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value not in (None, "")
    }


def _info_sheet_names(wb, requested_sheet: str) -> list[str]:
    if requested_sheet in wb.sheetnames:
        return [requested_sheet]
    if requested_sheet == INFO_SHEET:
        monthly = [name for name in wb.sheetnames if is_month_info_sheet(name)]
        if monthly:
            return monthly
    raise MissingSheetError(f"Không tìm thấy sheet “{requested_sheet}”.")


def read_sqt_items(daily_file: str, sheet_name: str = INFO_SHEET) -> List[SqtSelectionItem]:
    """Read unique SQT values in first-seen order and count source rows."""
    if not daily_file or not os.path.isfile(daily_file):
        raise DailyFileNotFoundError("Không tìm thấy file Excel hằng ngày.")

    try:
        if file_utils.is_file_locked(daily_file):
            raise DailyFileLockedError(
                "File Excel hằng ngày đang bị khóa. Vui lòng đóng Excel và thử lại."
            )
    except FileNotFoundError as exc:
        raise DailyFileNotFoundError("Không tìm thấy file Excel hằng ngày.") from exc

    wb = None
    try:
        wb = load_workbook(daily_file, read_only=True, data_only=True)
        sheet_names = _info_sheet_names(wb, sheet_name)

        counts: "OrderedDict[tuple[str, str], list[int]]" = OrderedDict()
        for current_sheet in sheet_names:
            ws = wb[current_sheet]
            headers = _header_map(ws)
            if SQT_COLUMN not in headers:
                raise MissingColumnError(
                    f"Không tìm thấy cột “{SQT_COLUMN}” trong sheet “{current_sheet}”."
                )

            sqt_col = headers[SQT_COLUMN]
            for row_number, row in enumerate(
                ws.iter_rows(
                    min_row=2,
                    min_col=sqt_col,
                    max_col=sqt_col,
                    values_only=True,
                ),
                start=2,
            ):
                sqt = normalize_sqt_value(row[0] if row else None)
                if not sqt:
                    continue
                key = (current_sheet if len(sheet_names) > 1 else "", sqt)
                counts.setdefault(key, []).append(row_number)

        if not counts:
            raise EmptySqtListError("Sheet không có SQT hợp lệ.")
        return [
            SqtSelectionItem(
                value=sqt,
                row_count=len(row_numbers),
                sheet_name=sheet,
                row_numbers=tuple(row_numbers),
            )
            for (sheet, sqt), row_numbers in counts.items()
        ]
    except SqtSelectionError:
        raise
    except PermissionError as exc:
        raise DailyFileLockedError(
            "File Excel hằng ngày đang bị khóa. Vui lòng đóng Excel và thử lại."
        ) from exc
    except (InvalidFileException, OSError, ValueError) as exc:
        raise SqtSelectionError(
            f"Không thể đọc file {os.path.basename(daily_file)}. "
            "Vui lòng đóng Excel và thử lại."
        ) from exc
    finally:
        if wb is not None:
            wb.close()


def write_selection_json(
    target_path: str | os.PathLike[str],
    daily_file: str,
    selected_sqt: Iterable[str],
    sheet_name: str = INFO_SHEET,
    *,
    operation: str = SELECTION_OPERATION,
    selected_items: Iterable[SqtSelectionItem] | None = None,
) -> str:
    """Write the selected SQT list atomically for PAD to consume."""
    selected = [str(value) for value in selected_sqt if str(value).strip()]
    if not selected:
        raise SelectionJsonWriteError("Vui lòng chọn ít nhất một Số quyết toán.")

    target = Path(target_path)
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise SelectionJsonWriteError("Không tạo được thư mục runtime.") from exc

    payload = {
        "operation": str(operation or SELECTION_OPERATION).strip() or SELECTION_OPERATION,
        "daily_file": str(Path(daily_file).resolve()),
        "sheet_name": sheet_name,
        "selected_sqt": selected,
    }
    item_payloads = [
        item.to_selection_payload()
        for item in (selected_items or [])
        if item.value in selected
    ]
    if item_payloads and operation == SELECTION_OPERATION:
        payload["mode"] = "monthly_info"
        payload["selected_items"] = item_payloads

    temp_path = target.with_name(target.name + ".tmp")
    try:
        with open(temp_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
            f.write("\n")
        os.replace(temp_path, target)
    except OSError as exc:
        try:
            if temp_path.exists():
                temp_path.unlink()
        except OSError:
            pass
        raise SelectionJsonWriteError("Không ghi được JSON lựa chọn RPA.") from exc

    return str(target.resolve())
