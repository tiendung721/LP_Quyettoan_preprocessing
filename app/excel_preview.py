"""Đọc và tạo dữ liệu xem trước (preview) cho file output.

Hỗ trợ Excel (.xlsx/.xlsm) bằng openpyxl và CSV bằng module csv.
Trả về danh sách sheet, dữ liệu 20 dòng đầu, số dòng và số cột.
"""

from __future__ import annotations

import csv
import os
from typing import Any, Dict, List

DEFAULT_MAX_ROWS = 20


def preview_file(path: str, max_rows: int = DEFAULT_MAX_ROWS) -> Dict[str, Any]:
    """Đọc file và trả về dữ liệu preview.

    Kết quả gồm các khóa:
        - sheets: danh sách tên sheet (CSV chỉ có 1 phần tử)
        - sheet_name: tên sheet đang xem
        - rows: list các dòng (mỗi dòng là list ô), tối đa ``max_rows`` dòng
        - row_count: tổng số dòng của sheet
        - column_count: tổng số cột của sheet
    """
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return _preview_csv(path, max_rows)
    return _preview_excel(path, max_rows)


# ---------------------------------------------------------------------------
def _preview_excel(path: str, max_rows: int) -> Dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        if not sheet_names:
            return {
                "sheets": [],
                "sheet_name": None,
                "rows": [],
                "row_count": 0,
                "column_count": 0,
            }

        first_name = sheet_names[0]
        ws = wb[first_name]

        rows: List[List[Any]] = []
        max_col_seen = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            values = ["" if v is None else v for v in row]
            max_col_seen = max(max_col_seen, len(values))
            rows.append(list(values))

        # Chuẩn hóa số cột cho tất cả các dòng preview.
        for r in rows:
            if len(r) < max_col_seen:
                r.extend([""] * (max_col_seen - len(r)))

        # Tổng số dòng/cột: ưu tiên dimension của openpyxl, nếu thiếu thì suy ra.
        row_count = ws.max_row if ws.max_row is not None else len(rows)
        column_count = ws.max_column if ws.max_column is not None else max_col_seen

        return {
            "sheets": sheet_names,
            "sheet_name": first_name,
            "rows": rows,
            "row_count": int(row_count or 0),
            "column_count": int(column_count or 0),
        }
    finally:
        wb.close()


# ---------------------------------------------------------------------------
def _preview_csv(path: str, max_rows: int) -> Dict[str, Any]:
    rows: List[List[Any]] = []
    total_rows = 0
    max_cols = 0

    # utf-8-sig để tự bỏ BOM nếu có; newline="" theo khuyến nghị module csv.
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            total_rows += 1
            max_cols = max(max_cols, len(row))
            if len(rows) < max_rows:
                rows.append(list(row))

    # Chuẩn hóa số cột cho các dòng preview.
    for r in rows:
        if len(r) < max_cols:
            r.extend([""] * (max_cols - len(r)))

    return {
        "sheets": ["CSV"],
        "sheet_name": "CSV",
        "rows": rows,
        "row_count": total_rows,
        "column_count": max_cols,
    }
