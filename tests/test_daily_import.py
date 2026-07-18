from __future__ import annotations

import json
import logging
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.daily_import import (
    EXPENSE_HEADERS,
    EXPENSE_SHEET,
    INFO_HEADERS,
    INFO_SHEET,
    REASON_MANY_BILLS,
    REASON_MANY_TARGETS,
    REASON_NO_MONTH,
    REASON_NO_TARGET,
    DailyImportError,
    DailyImportService,
    normalize_cargo,
    remove_extract_rows,
)
from app.database import Database


LEGACY_INFO_HEADERS = list(INFO_HEADERS)
LEGACY_EXPENSE_HEADERS = [name for name in EXPENSE_HEADERS if name != "Nơi đóng"]


class DailyImportServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.db = Database(str(self.root / "state.db"))
        self.db.init_db()
        self.service = DailyImportService(self.db, logging.getLogger("daily-test"))
        self.daily = self.root / "daily.xlsx"
        self.output = self.root / "output.json"
        self._make_daily()

    def tearDown(self) -> None:
        self.temp.cleanup()

    def _make_daily(self) -> None:
        wb = Workbook()
        info = wb.active
        info.title = INFO_SHEET
        info.append(LEGACY_INFO_HEADERS)
        values = {name: None for name in LEGACY_INFO_HEADERS}
        values.update(
            {
                "SQT PM": 100,
                "Ngày Đóng": datetime(2026, 6, 1),
                "Số Container": "ABCD1234567",
                "Số tấn": 28.0,
                "Loại hàng": "Vôi rời",
                "Nơi đóng": "Kho A",
                "Tên tàu": "TÀU CŨ 01S",
                "Ngày chạy": datetime(2026, 6, 2),
                "VT biển": "Hãng A",
                "Trạng thái nhập": "Đã nhập",
            }
        )
        info.append([values[name] for name in LEGACY_INFO_HEADERS])
        values.update(
            {
                "SQT PM": 101,
                "Ngày Đóng": datetime(2026, 6, 3),
                "Số Container": "EFGH1234567",
                "Số tấn": 27.0,
                "Loại hàng": "Vôi",
                "Nơi đóng": "Kho B",
                "Tên tàu": "TÀU B 02S",
                "Ngày chạy": datetime(2026, 6, 4),
                "VT biển": "Hãng B",
            }
        )
        info.append([values[name] for name in LEGACY_INFO_HEADERS])
        expense = wb.create_sheet(EXPENSE_SHEET)
        expense.append(LEGACY_EXPENSE_HEADERS)
        wb.save(self.daily)
        wb.close()

    def _make_monthly_daily(self, *, include_expense: bool = False) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tháng 7"
        ws.append(
            [
                "SQT PM",
                "Ngày Đóng",
                "Số Container",
                None,
                "Số chì",
                "Lô SX",
                "Số tấn",
                "Loại hàng",
                "Nơi đóng",
                "Tên tàu",
                "Ngày chạy",
                None,
                "Người nhận",
                "VT biển",
                "Vận chuyển",
                "HD HP",
                "HD HCM",
                "Ngày Giao",
                "Hóa Đơn",
                "Ghi chú",
            ]
        )
        ws.append(
            [
                716,
                datetime(2026, 7, 31),
                "JULY1234567",
                None,
                "SEAL716",
                "",
                28.5,
                "Vôi rời",
                "Kho tháng 7",
                "TÀU T7",
                datetime(2026, 8, 2),
                "DNA",
                "DNA",
                "Hãng T7",
                "PHB",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        if include_expense:
            expense = wb.create_sheet(EXPENSE_SHEET)
            expense.append(LEGACY_EXPENSE_HEADERS)
        wb.save(self.daily)
        wb.close()

    def _make_output(self) -> None:
        self._write_output_rows(
            {
                "file_nguon": "scale.jpg",
                "ma_md5_file": "scale-md5",
                "loai_chung_tu": "Phiếu cân",
                "trang_thai": "OK",
                "ngay_dong": "05/06/2026",
                "so_container": "IJKL 1234567",
                "so_tan": 29.5,
                "loai_hang": "VOI ROI",
                "noi_dong": "Kho C",
                "nguoi_nhan": "LÊ PHẠM",
                "truong_khac": {"so_phieu": "0000733"},
                "do_tin_cay": "Cao",
                "canh_bao": [],
            },
            {
                "file_nguon": "bill.pdf",
                "ma_md5_file": "bill-md5",
                "loai_chung_tu": "Bill",
                "trang_thai": "OK",
                "so_container": "IJKL1234567",
                "ten_tau": "TÀU C 03S",
                "ngay_chay": "06/06/2026",
                "vt_bien": "Hãng C",
                "so_bill": "BILL-03",
                "so_chi_seal": "SEAL03",
                "do_tin_cay": "Cao",
            },
            {
                "file_nguon": "cost.jpg",
                "ma_md5_file": "cost-md5",
                "loai_chung_tu": "Khoản chi",
                "trang_thai": "OK",
                "ngay_chay": "03/06/2026",
                "so_container": "EFGH1234567",
                "so_hd": "51",
                "don_gia": 1_700_000,
                "thanh_tien": 45_900_000,
                "vat": 3_672_000,
                "tong_tien": 49_572_000,
                "do_tin_cay": "Cao",
                "canh_bao": [{"noi_dung": "Kiểm tra VAT"}],
            },
        )

    def _write_output_rows(self, *rows: dict) -> None:
        payload = {
            "metadata": {
                "phien_ban_schema": "1.0",
                "tong_so_dong_boc_tach": len(rows),
            },
            "du_lieu_boc_tach": list(rows),
            "canh_bao": [],
            "raw_data": [],
        }
        with open(self.output, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    def test_normalize_cargo_with_vietnamese_accents(self) -> None:
        self.assertEqual(normalize_cargo("VOI")[0], "Vôi")
        self.assertEqual(normalize_cargo("VOI ROI")[0], "Vôi rời")
        self.assertEqual(normalize_cargo("Vôi cục C2 xá rời")[0], "Vôi cục C2 xá rời")

    def test_end_to_end_upgrade_match_and_deduplicate(self) -> None:
        self._make_output()
        analysis = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(analysis.extracted_rows, 3)
        self.assertEqual(analysis.new_info_count, 1)
        self.assertEqual(len(analysis.expense_changes), 1)
        self.assertEqual(analysis.expense_changes[0].values["SQT PM"], 101)

        summary = self.service.commit(analysis)
        self.assertEqual(summary.new_info, 1)
        self.assertEqual(summary.new_expenses, 1)
        # Không còn sao lưu file theo dõi: không tạo thư mục Backups.
        self.assertEqual(summary.backup_path, "")
        self.assertFalse((Path(self.daily).parent / "Backups").exists())

        wb = load_workbook(self.daily, data_only=False)
        info = wb[INFO_SHEET]
        expense = wb[EXPENSE_SHEET]
        self.assertIn("MD5", [cell.value for cell in info[1]])
        expense_headers = {cell.value: cell.column for cell in expense[1]}
        self.assertIn("MD5", expense_headers)
        self.assertIn("Nơi đóng", expense_headers)
        self.assertEqual(info.max_row, 4)
        self.assertEqual(expense.max_row, 2)
        headers = {cell.value: cell.column for cell in info[1]}
        self.assertEqual(info.cell(2, headers["Trạng thái nhập"]).value, "Đã nhập")
        self.assertEqual(info.cell(4, headers["SQT PM"]).value, 102)
        self.assertEqual(info.cell(4, headers["Loại hàng"]).value, "Vôi rời")
        self.assertEqual(info.cell(4, headers["Người nhận"]).value, "LÊ PHẠM")
        self.assertEqual(info.cell(4, headers["Trạng thái nhập"]).value, "Chưa nhập")
        self.assertEqual(expense.cell(2, expense_headers["Nơi đóng"]).value, "Kho B")
        wb.close()

        again = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(again.new_info_count, 0)
        self.assertEqual(len(again.expense_changes), 0)
        self.assertEqual(again.duplicate_documents, 3)

    def test_reimport_after_deleting_rows_from_daily(self) -> None:
        """File theo dõi là sổ cái: xóa dòng khỏi Excel thì nhập lại được ngay.

        Trước đây “đã nhập” chỉ nằm trong SQLite nên xóa dòng khỏi Excel xong nhập
        lại vẫn bị báo “đã nhập rồi”, phải xóa app_state.db mới nhập lại được.
        """
        self._make_output()
        self.service.commit(self.service.analyze(str(self.output), str(self.daily), 1))

        # Chưa xóa gì: chứng từ đã có MD5 trong file theo dõi -> bỏ qua.
        again = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(again.duplicate_documents, 3)
        self.assertFalse(again.has_changes)

        # Người dùng xóa các dòng app vừa ghi khỏi file theo dõi.
        wb = load_workbook(self.daily)
        wb[INFO_SHEET].delete_rows(4)  # dòng quyết toán mới (Phiếu cân + Bill)
        wb[EXPENSE_SHEET].delete_rows(2)  # dòng khoản chi
        wb.save(self.daily)
        wb.close()

        # MD5 không còn trong Excel -> nhập lại được, không phải đụng vào database.
        third = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(third.duplicate_documents, 0)
        self.assertEqual(third.new_info_count, 1)
        self.assertEqual(len(third.expense_changes), 1)
        self.assertEqual(third.expense_changes[0].values["SQT PM"], 101)

        summary = self.service.commit(third)
        self.assertEqual(summary.new_info, 1)
        self.assertEqual(summary.new_expenses, 1)

    def test_unwritten_rows_of_same_document_are_not_skipped(self) -> None:
        """Một file chứng từ sinh nhiều dòng: dòng chưa ghi không bị coi là đã nhập.

        Ảnh khoản chi của người dùng có 4 container dùng chung một MD5. Nếu chỉ đối
        chiếu theo MD5 thì 1 dòng được ghi sẽ làm 3 dòng còn lại bị bỏ qua oan.
        """
        shared = {
            "file_nguon": "chi-4-cont.jpg",
            "ma_md5_file": "cost-shared",
            "loai_chung_tu": "Khoản chi",
            "ngay_chay": "03/06/2026",
            "so_hd": "51",
        }
        self._write_output_rows(
            {**shared, "so_container": "EFGH1234567", "tong_tien": 1_000_000},
            {**shared, "so_container": "AAAA1111111", "tong_tien": 2_000_000},
            {**shared, "so_container": "BBBB2222222", "tong_tien": 3_000_000},
        )

        # Chỉ EFGH1234567 khớp SQT 101; hai dòng kia không ghép được.
        first = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(len(first.expense_changes), 1)
        self.assertEqual(first.expense_changes[0].values["Số Container"], "EFGH1234567")
        self.assertEqual(len(first.unmatched_rows), 2)

        # Người dùng bỏ hai dòng lỗi -> dòng đạt được ghi (MD5 vào file theo dõi).
        remove_extract_rows(
            str(self.output), [row.json_index for row in first.unmatched_rows]
        )
        second = self.service.analyze(str(self.output), str(self.daily), 1)
        self.service.commit(second)
        self.assertEqual(self._expense_sqt_column(), [101])

        # Tải lại đúng file JSON đó: dòng đã ghi bị bỏ qua, hai dòng kia vẫn được
        # xét lại chứ không bị "ăn theo" MD5 của dòng đã ghi.
        self._write_output_rows(
            {**shared, "so_container": "EFGH1234567", "tong_tien": 1_000_000},
            {**shared, "so_container": "AAAA1111111", "tong_tien": 2_000_000},
            {**shared, "so_container": "BBBB2222222", "tong_tien": 3_000_000},
        )
        third = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(third.expense_changes, [])
        self.assertEqual(
            sorted(row.container for row in third.unmatched_rows),
            ["AAAA1111111", "BBBB2222222"],
        )

    def test_bill_without_scale_is_persisted(self) -> None:
        self._write_output_rows(
            {
                "file_nguon": "waiting.pdf",
                "ma_md5_file": "waiting-bill",
                "loai_chung_tu": "Bill",
                "so_container": "MNOP1234567",
                "ten_tau": "TÀU CHỜ 01S",
                "ngay_chay": "10/06/2026",
                "vt_bien": "Hãng chờ",
            }
        )

        analysis = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(analysis.new_info_count, 1)
        self.assertEqual(analysis.info_changes[0].values["Số Container"], "MNOP1234567")
        self.assertEqual(analysis.info_changes[0].values["Tên tàu"], "TÀU CHỜ 01S")

        summary = self.service.commit(analysis)
        self.assertEqual(summary.new_info, 1)
        self.assertEqual(summary.pending_count, 0)

    def test_invalid_json_schema_is_reported(self) -> None:
        with open(self.output, "w", encoding="utf-8") as f:
            json.dump({"metadata": {}}, f)

        with self.assertRaises(DailyImportError) as ctx:
            self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertIn("du_lieu_boc_tach", str(ctx.exception))

    def test_multiple_bills_are_unmatched_and_can_be_dropped(self) -> None:
        rows = [
            {
                "file_nguon": "scale.jpg",
                "ma_md5_file": "scale-choice",
                "loai_chung_tu": "Phiếu cân",
                "trang_thai": "OK",
                "ngay_dong": "08/06/2026",
                "so_container": "QRST1234567",
                "so_tan": 28,
                "loai_hang": "VOI",
            }
        ]
        for md5, bill_no, vessel in (
            ("bill-choice-a", "BILL-A", "TÀU A 01S"),
            ("bill-choice-b", "BILL-B", "TÀU B 02S"),
        ):
            rows.append(
                {
                    "file_nguon": bill_no + ".pdf",
                    "ma_md5_file": md5,
                    "loai_chung_tu": "Bill",
                    "trang_thai": "OK",
                    "so_container": "QRST1234567",
                    "so_bill": bill_no,
                    "ten_tau": vessel,
                    "ngay_chay": "09/06/2026",
                    "vt_bien": "Hãng tàu",
                }
            )
        self._write_output_rows(*rows)

        # Container khớp nhiều Bill -> không tự đoán, không hỏi chọn tay.
        first = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(first.info_changes, [])
        self.assertEqual(len(first.unmatched_rows), 3)
        self.assertEqual(
            {row.reason for row in first.unmatched_rows}, {REASON_MANY_BILLS}
        )
        self.assertEqual(
            sorted(row.json_index for row in first.unmatched_rows), [0, 1, 2]
        )
        # Dòng lỗi không nằm lại hàng chờ để bắt người dùng xử lý sau.
        self.assertEqual(self.db.count_pending_rows(), 0)

        # Người dùng chọn “Hủy các dòng lỗi và nhập các dòng còn lại”.
        removed = remove_extract_rows(
            str(self.output), [row.json_index for row in first.unmatched_rows]
        )
        self.assertEqual(removed, 3)

        second = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(second.unmatched_rows, [])
        self.assertFalse(second.has_changes)
        with open(self.output, encoding="utf-8") as f:
            payload = json.load(f)
        self.assertEqual(payload["du_lieu_boc_tach"], [])
        self.assertEqual(payload["metadata"]["tong_so_dong_boc_tach"], 0)

    def test_expense_matches_by_date_and_container_without_invoice(self) -> None:
        self._write_output_rows(
            {
                "file_nguon": "cost-1.jpg",
                "ma_md5_file": "cost-one",
                "loai_chung_tu": "Khoản chi",
                "ngay_chay": "03/06/2026",
                "so_container": "EFGH1234567",
                "so_hd": "OLD",
                "don_gia": 100,
            }
        )
        first = self.service.analyze(str(self.output), str(self.daily), 1)
        self.service.commit(first)

        self._write_output_rows(
            {
                "file_nguon": "cost-2.jpg",
                "ma_md5_file": "cost-two",
                "loai_chung_tu": "Khoản chi",
                "ngay_chay": "03/06/2026",
                "so_container": "EFGH1234567",
                "so_hd": "NEW",
                "vat": 10,
            }
        )
        second = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(len(second.expense_changes), 1)
        self.assertEqual(second.expense_changes[0].action, "UPDATE")
        self.assertEqual(second.expense_changes[0].values["Nơi đóng"], "Kho B")

    def test_expense_place_follows_updated_settlement_row(self) -> None:
        self._write_output_rows(
            {
                "file_nguon": "cost-place.jpg",
                "ma_md5_file": "cost-place",
                "loai_chung_tu": "Khoản chi",
                "ngay_chay": "03/06/2026",
                "so_container": "EFGH1234567",
                "so_hd": "PLACE",
                "don_gia": 100,
            }
        )
        self.service.commit(self.service.analyze(str(self.output), str(self.daily), 1))
        self.assertEqual(self._expense_place_column(), ["Kho B"])

        self._write_output_rows(
            {
                "file_nguon": "scale-place-update.jpg",
                "ma_md5_file": "scale-place-update",
                "loai_chung_tu": "Phiếu cân",
                "trang_thai": "OK",
                "ngay_dong": "03/06/2026",
                "so_container": "EFGH1234567",
                "so_tan": 27.0,
                "loai_hang": "Vôi",
                "noi_dong": "Kho D",
                "do_tin_cay": "Cao",
            }
        )
        analysis = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(len(analysis.info_changes), 1)
        self.assertEqual(analysis.expense_changes, [])
        self.assertIn("Nơi đóng", {conflict.field_name for conflict in analysis.conflicts})

        decisions = {conflict.conflict_id: True for conflict in analysis.conflicts}
        self.service.commit(analysis, conflict_decisions=decisions)

        self.assertEqual(self._expense_place_column(), ["Kho D"])

    def test_missing_md5_still_imports(self) -> None:
        self._write_output_rows(
            {
                "file_nguon": "scale-no-md5.jpg",
                "loai_chung_tu": "Phiếu cân",
                "ngay_dong": "11/06/2026",
                "so_container": "ZZZZ1234567",
                "bien_so_xe": "15H 22404",
                "so_tan": 25,
                "loai_hang": "VOI",
            }
        )

        analysis = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(analysis.new_info_count, 1)
        self.assertEqual(analysis.info_changes[0].values["Biển số xe"], "15H 22404")
        self.assertEqual(analysis.info_changes[0].values["MD5"], "")

    def test_scale_matches_existing_row_by_date_and_container_ignoring_cargo(self) -> None:
        self._write_output_rows(
            {
                "file_nguon": "scale-cargo-mismatch.jpg",
                "ma_md5_file": "scale-cargo-mismatch",
                "loai_chung_tu": "Phiếu cân",
                "trang_thai": "OK",
                "ngay_dong": "01/06/2026",
                "so_container": "ABCD1234567",
                "bien_so_xe": "15H-15912",
                "so_tan": 27.88,
                "loai_hang": "VOI",
                "noi_dong": "Kho A",
                "do_tin_cay": "Cao",
            }
        )

        analysis = self.service.analyze(str(self.output), str(self.daily), 1)

        self.assertEqual(analysis.new_info_count, 0)
        self.assertEqual(analysis.updated_info_count, 1)
        change = analysis.info_changes[0]
        self.assertEqual(change.action, "UPDATE")
        self.assertEqual(change.target_row, 2)
        self.assertEqual(change.sqt, 100)
        self.assertEqual(change.values["SQT PM"], 100)
        self.assertEqual(change.values["Số Container"], "ABCD1234567")
        self.assertIn("Loại hàng", {conflict.field_name for conflict in analysis.conflicts})

    def test_scale_matching_multiple_sqt_is_unmatched(self) -> None:
        self._append_info_row(200, datetime(2026, 6, 12), "MULT1234567", 20)
        self._append_info_row(201, datetime(2026, 6, 12), "MULT1234567", 21)

        self._write_output_rows(
            {
                "file_nguon": "scale-multi.jpg",
                "ma_md5_file": "scale-multi",
                "loai_chung_tu": "Phiếu cân",
                "ngay_dong": "12/06/2026",
                "so_container": "MULT1234567",
                "so_tan": 22,
            }
        )
        analysis = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(analysis.info_changes, [])
        self.assertEqual(len(analysis.unmatched_rows), 1)
        self.assertEqual(analysis.unmatched_rows[0].reason, REASON_MANY_TARGETS)
        self.assertEqual(analysis.unmatched_rows[0].doc_label, "Phiếu cân")

    def test_bill_of_ambiguous_container_is_reported_not_dropped(self) -> None:
        """Phiếu cân khớp nhiều SQT: Bill đi kèm cũng phải được báo, không âm thầm bỏ."""
        self._append_info_row(200, datetime(2026, 6, 12), "MULT1234567", 20)
        self._append_info_row(201, datetime(2026, 6, 12), "MULT1234567", 21)

        self._write_output_rows(
            {
                "file_nguon": "scale-multi.jpg",
                "ma_md5_file": "scale-multi",
                "loai_chung_tu": "Phiếu cân",
                "ngay_dong": "12/06/2026",
                "so_container": "MULT1234567",
                "so_tan": 22,
            },
            {
                "file_nguon": "bill-multi.pdf",
                "ma_md5_file": "bill-multi",
                "loai_chung_tu": "Bill",
                "so_container": "MULT1234567",
                "so_bill": "BL-MULTI",
                "ten_tau": "TÀU M 09S",
                "ngay_chay": "13/06/2026",
            },
        )

        analysis = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(analysis.info_changes, [])
        self.assertEqual(
            sorted(row.doc_label for row in analysis.unmatched_rows),
            ["Bill", "Phiếu cân"],
        )
        self.assertEqual(
            {row.reason for row in analysis.unmatched_rows}, {REASON_MANY_TARGETS}
        )
        self.assertEqual(
            sorted(row.json_index for row in analysis.unmatched_rows), [0, 1]
        )

    def test_expense_without_matching_sqt_is_not_written(self) -> None:
        self._write_output_rows(
            {
                "file_nguon": "cost-orphan.jpg",
                "ma_md5_file": "cost-orphan",
                "loai_chung_tu": "Khoản chi",
                "ngay_chay": "20/06/2026",
                "so_container": "WXYZ1234567",
                "so_hd": "77",
                "don_gia": 1_500_000,
                "tong_tien": 42_000_000,
            }
        )

        analysis = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(analysis.expense_changes, [])
        self.assertEqual(len(analysis.unmatched_rows), 1)
        unmatched = analysis.unmatched_rows[0]
        self.assertEqual(unmatched.doc_label, "Khoản chi")
        self.assertEqual(unmatched.container, "WXYZ1234567")
        self.assertEqual(unmatched.match_date, "2026-06-20")
        self.assertEqual(unmatched.reference, "HĐ 77")
        self.assertEqual(unmatched.reason, REASON_NO_TARGET)
        self.assertEqual(unmatched.json_index, 0)
        self.assertEqual(self.db.count_pending_rows(), 0)

        self.service.commit(analysis)
        self.assertEqual(self._expense_sqt_column(), [])

    def test_expense_matching_multiple_sqt_is_not_written(self) -> None:
        # Cùng container + cùng ngày nhưng hai dòng quyết toán -> không rõ SQT nào.
        self._append_info_row(300, datetime(2026, 6, 3), "EFGH1234567", 27)

        self._write_output_rows(
            {
                "file_nguon": "cost-ambiguous.jpg",
                "ma_md5_file": "cost-ambiguous",
                "loai_chung_tu": "Khoản chi",
                "ngay_chay": "03/06/2026",
                "so_container": "EFGH1234567",
                "so_hd": "88",
                "tong_tien": 10_000_000,
            }
        )

        analysis = self.service.analyze(str(self.output), str(self.daily), 1)
        self.assertEqual(analysis.expense_changes, [])
        self.assertEqual(len(analysis.unmatched_rows), 1)
        self.assertEqual(analysis.unmatched_rows[0].reason, REASON_MANY_TARGETS)

        self.service.commit(analysis)
        self.assertEqual(self._expense_sqt_column(), [])

    def test_monthly_workbook_creates_next_month_sheet_and_global_sqt(self) -> None:
        self._make_monthly_daily()
        self._write_output_rows(
            {
                "file_nguon": "scale-aug.jpg",
                "ma_md5_file": "scale-aug",
                "loai_chung_tu": "Phiếu cân",
                "ngay_dong": "05/08/2026",
                "so_container": "AUGU1234567",
                "so_tan": 29.5,
                "loai_hang": "VOI ROI",
                "noi_dong": "Kho tháng 8",
                "do_tin_cay": "Cao",
                "canh_bao": [],
            }
        )

        analysis = self.service.analyze(str(self.output), str(self.daily), 1)

        self.assertEqual(analysis.new_info_count, 1)
        change = analysis.info_changes[0]
        self.assertEqual(change.sqt, 717)
        self.assertEqual(change.target_sheet, "Tháng 8")

        self.service.commit(analysis)
        wb = load_workbook(self.daily)
        self.assertIn("Tháng 8", wb.sheetnames)
        self.assertIn(EXPENSE_SHEET, wb.sheetnames)
        ws = wb["Tháng 8"]
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        self.assertIn("Số chì/Seal", headers)
        self.assertIn("Hóa Đơn quyết toán", headers)
        self.assertEqual(ws.cell(2, headers["SQT PM"]).value, 717)
        self.assertEqual(ws.cell(2, headers["Số Container"]).value, "AUGU1234567")
        self.assertTrue(
            ws.column_dimensions[
                ws.cell(1, headers["MD5"]).column_letter
            ].hidden
        )
        wb.close()

    def test_monthly_workbook_missing_close_date_is_unmatched(self) -> None:
        self._make_monthly_daily()
        self._write_output_rows(
            {
                "file_nguon": "scale-no-date.jpg",
                "ma_md5_file": "scale-no-date",
                "loai_chung_tu": "Phiếu cân",
                "so_container": "NODT1234567",
                "so_tan": 29.5,
                "loai_hang": "VOI ROI",
            }
        )

        analysis = self.service.analyze(str(self.output), str(self.daily), 1)

        self.assertEqual(analysis.info_changes, [])
        self.assertEqual(len(analysis.unmatched_rows), 1)
        self.assertEqual(analysis.unmatched_rows[0].reason, REASON_NO_MONTH)

    def test_expense_matches_info_from_monthly_sheets(self) -> None:
        self._make_monthly_daily()
        self._write_output_rows(
            {
                "file_nguon": "cost-monthly.jpg",
                "ma_md5_file": "cost-monthly",
                "loai_chung_tu": "Khoản chi",
                "ngay_chay": "31/07/2026",
                "so_container": "JULY1234567",
                "so_hd": "T7",
                "don_gia": 100,
            }
        )

        analysis = self.service.analyze(str(self.output), str(self.daily), 1)

        self.assertEqual(len(analysis.expense_changes), 1)
        self.assertEqual(analysis.expense_changes[0].values["SQT PM"], 716)
        self.assertEqual(analysis.expense_changes[0].values["Nơi đóng"], "Kho tháng 7")
        self.service.commit(analysis)
        self.assertEqual(self._expense_sqt_column(), [716])

    # ------------------------------------------------------------------ #
    def _append_info_row(
        self, sqt: int, close_date: datetime, container: str, tons: float
    ) -> None:
        wb = load_workbook(self.daily)
        ws = wb[INFO_SHEET]
        values = {name: None for name in LEGACY_INFO_HEADERS}
        values.update(
            {
                "SQT PM": sqt,
                "Ngày Đóng": close_date,
                "Số Container": container,
                "Số tấn": tons,
            }
        )
        ws.append([values[name] for name in LEGACY_INFO_HEADERS])
        wb.save(self.daily)
        wb.close()

    def _expense_sqt_column(self) -> list:
        """Các giá trị SQT PM đã ghi vào sheet Khoan_Chi."""
        wb = load_workbook(self.daily)
        expense = wb[EXPENSE_SHEET]
        headers = {cell.value: cell.column for cell in expense[1]}
        written = [
            expense.cell(row, headers["SQT PM"]).value
            for row in range(2, expense.max_row + 1)
        ]
        wb.close()
        return written

    def _expense_place_column(self) -> list:
        """Các giá trị Nơi đóng đã ghi vào sheet Khoan_Chi."""
        wb = load_workbook(self.daily)
        expense = wb[EXPENSE_SHEET]
        headers = {cell.value: cell.column for cell in expense[1]}
        written = [
            expense.cell(row, headers["Nơi đóng"]).value
            for row in range(2, expense.max_row + 1)
        ]
        wb.close()
        return written


if __name__ == "__main__":
    unittest.main()
