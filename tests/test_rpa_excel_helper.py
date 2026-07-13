from __future__ import annotations

import importlib.util
import json
import subprocess
import sys
import tempfile
import time
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
HELPER_PATH = ROOT / "scripts" / "rpa_excel_helper.py"
spec = importlib.util.spec_from_file_location("rpa_excel_helper", HELPER_PATH)
helper = importlib.util.module_from_spec(spec)
assert spec.loader is not None
spec.loader.exec_module(helper)


class RpaExcelHelperTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.daily = self.root / "quyet_toan_hang_ngay.xlsx"
        self._make_workbook(include_last_import=True)

    def tearDown(self) -> None:
        self.temp.cleanup()

    def _make_workbook(self, *, include_last_import: bool) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Thong_Tin_Quyet_Toan"
        headers = ["SQT PM", "Số Container", "Trạng thái nhập"]
        if include_last_import:
            headers.append("Ngày nhập cuối")
        headers.extend(["Ngày cập nhật", "Ghi chú"])
        ws.append(headers)
        ws.append(["596", "CONT01", "Chưa nhập", "", "01/07/2026 08:00:00", "row 2"])
        ws.append(["596", "CONT02", "Đã nhập", "01/07/2026 09:00:00", "02/07/2026 08:00:00", "row 3"])
        ws.append(["597", "CONT03", "", "", "03/07/2026 08:00:00", "row 4"])
        other = wb.create_sheet("Sheet_Khac")
        other["A1"] = "Giữ nguyên"
        other["B1"] = "=1+1"
        wb.save(self.daily)
        wb.close()

    def _headers(self, ws) -> dict[str, int]:
        return {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}

    def test_update_row_with_chua_nhap_status(self) -> None:
        result = helper.update_information_status(
            self.daily,
            "Thong_Tin_Quyet_Toan",
            2,
        )

        wb = load_workbook(self.daily, data_only=False)
        ws = wb["Thong_Tin_Quyet_Toan"]
        headers = self._headers(ws)
        self.assertTrue(result["success"])
        self.assertEqual(result["sqt"], "596")
        self.assertEqual(ws.cell(2, headers["Trạng thái nhập"]).value, "Đã nhập")
        self.assertEqual(
            ws.cell(2, headers["Ngày nhập cuối"]).value,
            result["updated_at"],
        )
        self.assertEqual(ws.cell(2, headers["Ngày cập nhật"]).value, "01/07/2026 08:00:00")
        self.assertEqual(ws.cell(3, headers["Trạng thái nhập"]).value, "Đã nhập")
        self.assertEqual(ws.cell(3, headers["Ngày nhập cuối"]).value, "01/07/2026 09:00:00")
        self.assertEqual(wb["Sheet_Khac"]["B1"].value, "=1+1")
        self.assertNotIn("RPA_Queue", wb.sheetnames)
        wb.close()

    def test_update_existing_imported_row_is_idempotent(self) -> None:
        first = helper.update_information_status(self.daily, "Thong_Tin_Quyet_Toan", 3)
        time.sleep(1.1)
        second = helper.update_information_status(self.daily, "Thong_Tin_Quyet_Toan", 3)

        wb = load_workbook(self.daily)
        ws = wb["Thong_Tin_Quyet_Toan"]
        headers = self._headers(ws)
        self.assertEqual(ws.cell(3, headers["Trạng thái nhập"]).value, "Đã nhập")
        self.assertEqual(ws.cell(3, headers["Ngày nhập cuối"]).value, second["updated_at"])
        self.assertNotEqual(first["updated_at"], second["updated_at"])
        wb.close()

    def test_only_requested_row_is_updated_for_same_sqt(self) -> None:
        helper.update_information_status(self.daily, "Thong_Tin_Quyet_Toan", 2)

        wb = load_workbook(self.daily)
        ws = wb["Thong_Tin_Quyet_Toan"]
        headers = self._headers(ws)
        self.assertEqual(ws.cell(2, headers["Ngày nhập cuối"]).value is not None, True)
        self.assertEqual(ws.cell(3, headers["Ngày nhập cuối"]).value, "01/07/2026 09:00:00")
        wb.close()

    def test_missing_last_import_column_is_added_once(self) -> None:
        self._make_workbook(include_last_import=False)

        helper.update_information_status(self.daily, "Thong_Tin_Quyet_Toan", 2)
        helper.update_information_status(self.daily, "Thong_Tin_Quyet_Toan", 2)

        wb = load_workbook(self.daily)
        ws = wb["Thong_Tin_Quyet_Toan"]
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers.count("Ngày nhập cuối"), 1)
        wb.close()

    def test_missing_file_sheet_and_required_column_errors(self) -> None:
        with self.assertRaises(helper.ExcelHelperError) as missing_file:
            helper.update_information_status(self.root / "missing.xlsx", "Thong_Tin_Quyet_Toan", 2)
        self.assertEqual(missing_file.exception.exit_code, helper.ExitCode.FILE_NOT_FOUND)

        with self.assertRaises(helper.ExcelHelperError) as missing_sheet:
            helper.update_information_status(self.daily, "Sheet_Khong_Co", 2)
        self.assertEqual(missing_sheet.exception.exit_code, helper.ExitCode.SHEET_NOT_FOUND)

        wb = load_workbook(self.daily)
        ws = wb["Thong_Tin_Quyet_Toan"]
        headers = self._headers(ws)
        ws.cell(1, headers["Trạng thái nhập"]).value = "Trang thai khac"
        wb.save(self.daily)
        wb.close()
        with self.assertRaises(helper.ExcelHelperError) as missing_column:
            helper.update_information_status(self.daily, "Thong_Tin_Quyet_Toan", 2)
        self.assertEqual(
            missing_column.exception.exit_code,
            helper.ExitCode.REQUIRED_COLUMN_NOT_FOUND,
        )

    def test_invalid_rows_and_empty_row_are_rejected(self) -> None:
        for row in (1, 0, -5, 99):
            with self.assertRaises(helper.ExcelHelperError) as ctx:
                helper.update_information_status(self.daily, "Thong_Tin_Quyet_Toan", row)
            self.assertEqual(ctx.exception.exit_code, helper.ExitCode.INVALID_ROW)

        wb = load_workbook(self.daily)
        ws = wb["Thong_Tin_Quyet_Toan"]
        ws.append([None, None, None, None, None, None])
        empty_row = ws.max_row
        wb.save(self.daily)
        wb.close()

        with self.assertRaises(helper.ExcelHelperError) as empty:
            helper.update_information_status(self.daily, "Thong_Tin_Quyet_Toan", empty_row)
        self.assertEqual(empty.exception.exit_code, helper.ExitCode.INVALID_ROW)

    def test_cli_success_outputs_one_line_json(self) -> None:
        proc = subprocess.run(
            [
                sys.executable,
                str(HELPER_PATH),
                "finish-info",
                "--file",
                str(self.daily),
                "--sheet",
                "Thong_Tin_Quyet_Toan",
                "--row",
                "2",
            ],
            cwd=ROOT,
            text=True,
            encoding="utf-8",
            capture_output=True,
        )

        self.assertEqual(proc.returncode, 0, proc.stderr)
        self.assertEqual(proc.stderr, "")
        lines = proc.stdout.strip().splitlines()
        self.assertEqual(len(lines), 1)
        payload = json.loads(lines[0])
        self.assertTrue(payload["success"])
        self.assertEqual(payload["command"], "finish-info")
        self.assertEqual(payload["row"], 2)
        self.assertEqual(payload["status"], "Đã nhập")

    def test_cli_error_outputs_stderr_json_and_non_zero_exit(self) -> None:
        proc = subprocess.run(
            [
                sys.executable,
                str(HELPER_PATH),
                "finish-info",
                "--file",
                str(self.daily),
                "--sheet",
                "Thong_Tin_Quyet_Toan",
                "--row",
                "1",
            ],
            cwd=ROOT,
            text=True,
            encoding="utf-8",
            capture_output=True,
        )

        self.assertEqual(proc.returncode, int(helper.ExitCode.INVALID_ROW))
        self.assertEqual(proc.stdout, "")
        payload = json.loads(proc.stderr.strip())
        self.assertFalse(payload["success"])
        self.assertEqual(payload["error_code"], "INVALID_ROW")

    def test_cli_invalid_arguments_return_exit_code_2(self) -> None:
        proc = subprocess.run(
            [sys.executable, str(HELPER_PATH), "finish-info"],
            cwd=ROOT,
            text=True,
            encoding="utf-8",
            capture_output=True,
        )

        self.assertEqual(proc.returncode, int(helper.ExitCode.INVALID_ARGUMENTS))
        self.assertEqual(proc.stdout, "")
        payload = json.loads(proc.stderr.strip())
        self.assertEqual(payload["error_code"], "INVALID_ARGUMENTS")


if __name__ == "__main__":
    unittest.main()
