from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.daily_import import EXPENSE_SHEET, INFO_SHEET
from app.services.sqt_selection_service import (
    EmptySqtListError,
    EXPENSE_SELECTION_OPERATION,
    MissingColumnError,
    MissingSheetError,
    MultipleSheetsSelectedError,
    normalize_sqt_value,
    read_sqt_items,
    resolve_selection_sheet,
    write_selection_json,
)


class SqtSelectionServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.daily = self.root / "Quyết toán hằng ngày.xlsx"

    def tearDown(self) -> None:
        self.temp.cleanup()

    def _make_daily(self, values, *, sheet_name: str = INFO_SHEET, header: str = "SQT PM") -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append([header, "Số Container"])
        for index, value in enumerate(values, start=1):
            ws.append([value, f"CONT{index:02d}"])
        wb.save(self.daily)
        wb.close()

    def _make_monthly_daily(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tháng 7"
        ws.append(["SQT PM", "Số Container"])
        ws.append([716, "JULY1234567"])
        ws.append([716, "JULY7654321"])
        ws2 = wb.create_sheet("Tháng 8")
        ws2.append(["SQT PM", "Số Container"])
        ws2.append([717, "AUGU1234567"])
        wb.save(self.daily)
        wb.close()

    def test_normalize_sqt_value(self) -> None:
        self.assertEqual(normalize_sqt_value(596), "596")
        self.assertEqual(normalize_sqt_value(596.0), "596")
        self.assertEqual(normalize_sqt_value("596"), "596")
        self.assertEqual(normalize_sqt_value("596.0"), "596")
        self.assertEqual(normalize_sqt_value(""), "")
        self.assertEqual(normalize_sqt_value(None), "")

    def test_read_sqt_items_keeps_first_seen_order_and_counts_rows(self) -> None:
        self._make_daily([596, 596.0, "597", "", None, "596", 598.0])

        items = read_sqt_items(str(self.daily))

        self.assertEqual(
            [(item.value, item.row_count, item.display_text) for item in items],
            [
                ("596", 3, "596 — 3 dòng"),
                ("597", 1, "597 — 1 dòng"),
                ("598", 1, "598 — 1 dòng"),
            ],
        )

    def test_missing_sheet_is_reported(self) -> None:
        self._make_daily([596], sheet_name="SheetKhac")

        with self.assertRaises(MissingSheetError):
            read_sqt_items(str(self.daily))

    def test_missing_sqt_column_is_reported(self) -> None:
        self._make_daily([596], header="SQT Khac")

        with self.assertRaises(MissingColumnError):
            read_sqt_items(str(self.daily))

    def test_empty_sqt_list_is_reported(self) -> None:
        self._make_daily(["", None])

        with self.assertRaises(EmptySqtListError):
            read_sqt_items(str(self.daily))

    def test_write_selection_json_schema_and_overwrite(self) -> None:
        target = self.root / "runtime" / "rpa_input_selection.json"

        first = write_selection_json(target, str(self.daily), ["596"])
        second = write_selection_json(target, str(self.daily), ["597", "598"])

        self.assertEqual(first, second)
        with open(target, encoding="utf-8") as f:
            payload = json.load(f)

        self.assertEqual(payload["operation"], "NHAP_THONG_TIN")
        self.assertEqual(Path(payload["daily_file"]), self.daily.resolve())
        self.assertEqual(payload["sheet_name"], INFO_SHEET)
        self.assertEqual(payload["selected_sqt"], ["597", "598"])
        self.assertNotIn("RPA_Queue", json.dumps(payload, ensure_ascii=False))
        self.assertNotIn("CREATE_NEW", json.dumps(payload, ensure_ascii=False))
        self.assertFalse(target.with_name(target.name + ".tmp").exists())

    def test_read_monthly_info_items_and_write_selected_items(self) -> None:
        self._make_monthly_daily()
        target = self.root / "runtime" / "rpa_input_selection.json"

        items = read_sqt_items(str(self.daily))
        write_selection_json(
            target,
            str(self.daily),
            ["716", "717"],
            selected_items=items,
        )

        self.assertEqual(
            [
                (item.value, item.row_count, item.sheet_name, item.row_numbers)
                for item in items
            ],
            [
                ("716", 2, "Tháng 7", (2, 3)),
                ("717", 1, "Tháng 8", (2,)),
            ],
        )
        with open(target, encoding="utf-8") as f:
            payload = json.load(f)
        self.assertEqual(payload["mode"], "monthly_info")
        self.assertEqual(
            payload["selected_items"],
            [
                {"sqt": "716", "sheet_name": "Tháng 7", "row_numbers": [2, 3]},
                {"sqt": "717", "sheet_name": "Tháng 8", "row_numbers": [2]},
            ],
        )

    def test_resolve_selection_sheet_uses_month_sheet(self) -> None:
        self._make_monthly_daily()
        items = read_sqt_items(str(self.daily))
        july = [item for item in items if item.value == "716"]

        self.assertEqual(resolve_selection_sheet(INFO_SHEET, july), "Tháng 7")

    def test_resolve_selection_sheet_falls_back_to_default(self) -> None:
        self._make_daily([596, 597])
        items = read_sqt_items(str(self.daily))

        self.assertEqual(resolve_selection_sheet(INFO_SHEET, items), INFO_SHEET)
        self.assertEqual(resolve_selection_sheet(INFO_SHEET, None), INFO_SHEET)

    def test_resolve_selection_sheet_rejects_multiple_months(self) -> None:
        self._make_monthly_daily()
        items = read_sqt_items(str(self.daily))

        with self.assertRaises(MultipleSheetsSelectedError) as ctx:
            resolve_selection_sheet(INFO_SHEET, items)
        self.assertEqual(ctx.exception.sheets, ["Tháng 7", "Tháng 8"])

    def test_write_selection_json_supports_expense_operation(self) -> None:
        target = self.root / "runtime" / "rpa_input_selection.json"

        write_selection_json(
            target,
            str(self.daily),
            ["700"],
            sheet_name=EXPENSE_SHEET,
            operation=EXPENSE_SELECTION_OPERATION,
        )

        with open(target, encoding="utf-8") as f:
            payload = json.load(f)

        self.assertEqual(payload["operation"], "NHAP_KHOAN_CHI")
        self.assertEqual(payload["sheet_name"], EXPENSE_SHEET)
        self.assertEqual(payload["selected_sqt"], ["700"])


if __name__ == "__main__":
    unittest.main()
